import csv
from django.views.decorators.csrf import csrf_exempt
import calendar
import csv
import json
import traceback
from datetime import datetime, date, timedelta

import openpyxl
from allauth.account.models import EmailAddress
from allauth.socialaccount.models import SocialAccount
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.core.cache import cache
from django.core.mail import send_mail
from django.core.management import call_command
from django.db import IntegrityError, transaction
from django.db.models import Sum, Q
from django.db.models.functions import TruncMonth, TruncDay
from django.forms import modelformset_factory
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy, reverse
from django.utils import timezone
from django.utils.html import mark_safe, format_html, format_html_join
from django.views import generic
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import (
    TemplateView,
    ListView,
    CreateView,
    UpdateView,
    DeleteView,
    DetailView,
    View,
)

from finance_tracker.ai_utils import predict_category_ai
from .forms import (
    ExpenseForm,
    IncomeForm,
    RecurringTransactionForm,
    ProfileUpdateForm,
    CustomSignupForm,
    ContactForm,
    CashCreditForm,
    CashCreditRepaymentForm,
)
from .services import scan_bill_image
from .models import (
    CreditCard,
    Expense,
    Category,
    Income,
    PaymentSource,
    RecurringTransaction,
    UserProfile,
    SubscriptionPlan,
    Friend,
    CashCredit,
    CashCreditRepayment,
)
from .models import Notification


def create_category_ajax(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            name = data.get('name', '').strip()
            
            if not name:
                return JsonResponse({'success': False, 'error': 'Category name cannot be empty.'}, status=400)
            
            # Check Limits
            current_count = Category.objects.filter(user=request.user).count()
            limit = 20 # Free
            if request.user.profile.is_plus:
                limit = 10
            if request.user.profile.is_pro:
                limit = float('inf')

            if current_count >= limit:
                 return JsonResponse({'success': False, 'error': f'Category limit reached ({limit}). Please upgrade.'}, status=403)

            category = Category.objects.create(user=request.user, name=name)
            return JsonResponse({'success': True, 'id': category.id, 'name': category.name})
            
        except IntegrityError:
            return JsonResponse({'success': False, 'error': 'This category already exists.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
            
    return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=405)


@login_required
def get_payment_sources_ajax(request):
    """
    AJAX endpoint to get payment sources based on payment method and amount.
    Returns filtered list with balance information.
    """
    if request.method == 'GET':
        try:
            from .models import PaymentSource, CreditCard
            from decimal import Decimal
            
            payment_method = request.GET.get('payment_method', '')
            amount_str = request.GET.get('amount', '0')
            
            try:
                amount = Decimal(amount_str) if amount_str else Decimal('0')
            except:
                amount = Decimal('0')
            
            sources = []
            
            if payment_method == 'Cash':
                # Cash doesn't need a payment source
                return JsonResponse({
                    'success': True,
                    'sources': [],
                    'message': 'Cash payments do not require a payment source selection'
                })
            
            elif payment_method == 'Credit Card':
                # Return only credit cards
                credit_cards = CreditCard.objects.filter(
                    user=request.user,
                    is_active=True
                ).order_by('bank_name', 'name')
                
                for card in credit_cards:
                    has_sufficient_funds = card.available_limit >= amount if amount > 0 else True
                    sources.append({
                        'id': f'card_{card.id}',
                        'name': card.name,
                        'bank': card.bank_name,
                        'balance': float(card.available_limit),
                        'total_limit': float(card.credit_limit),
                        'display': f"{card.name} ({card.bank_name})",
                        'balance_display': f"₹{card.available_limit:,.2f} / ₹{card.credit_limit:,.2f}",
                        'type': 'credit_card',
                        'sufficient_funds': has_sufficient_funds,
                        'disabled': not has_sufficient_funds
                    })
            
            else:
                # For other payment methods (Debit Card, UPI, NetBanking), return payment sources
                payment_sources = PaymentSource.objects.filter(
                    user=request.user,
                    is_active=True
                ).order_by('account_type', 'name')
                
                for source in payment_sources:
                    has_sufficient_funds = source.balance >= amount if amount > 0 else True
                    sources.append({
                        'id': f'source_{source.id}',
                        'name': source.name,
                        'account_type': source.get_account_type_display(),
                        'bank': source.bank_name or '',
                        'balance': float(source.balance),
                        'display': source.name,
                        'balance_display': f"₹{source.balance:,.2f}",
                        'type': 'payment_source',
                        'sufficient_funds': has_sufficient_funds,
                        'disabled': not has_sufficient_funds
                    })
            
            return JsonResponse({
                'success': True,
                'sources': sources,
                'payment_method': payment_method,
                'amount': float(amount)
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=405)


@login_required
def scan_bill_ajax(request):
    """
    Scan uploaded bill image via Gemini and return extracted fields for form autofill.
    """
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Invalid request method."}, status=405)

    bill_file = request.FILES.get("bill_image")
    if not bill_file:
        return JsonResponse({"success": False, "error": "No bill image uploaded."}, status=400)

    if bill_file.size > 5 * 1024 * 1024:
        return JsonResponse({"success": False, "error": "Image too large. Max size is 5MB."}, status=400)

    allowed_mime_types = {"image/jpeg", "image/jpg", "image/png", "image/webp"}
    mime_type = (bill_file.content_type or "").lower()
    if mime_type not in allowed_mime_types:
        return JsonResponse(
            {
                "success": False,
                "error": "Unsupported file type. Please upload JPG, PNG, or WEBP image.",
            },
            status=400,
        )

    # get name and id of the categories
    categories = Category.objects.filter(user=request.user).order_by('name').values('name', 'id')

    result = scan_bill_image(image_bytes=bill_file.read(), mime_type=mime_type, categories=categories)
    status_code = 200 if result.get("success") else 400
    return JsonResponse(result, status=status_code)


def resend_verification_email(request):
    """
    AJAX view to resend verification email.
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email = data.get('email')
            
            # If email is not in body, try to get from logged in user
            if not email and request.user.is_authenticated:
                email = request.user.email
            
            # Fallback: Check allauth session key
            if not email:
                email = request.session.get('account_email')

            if not email:
                return JsonResponse({'success': False, 'error': 'Email is missing.'}, status=400)
            
            try:
                # Case-insensitive lookup just in case
                email_address = EmailAddress.objects.filter(email__iexact=email).first()
                if not email_address:
                     return JsonResponse({'success': False, 'error': f'Email {email} not found in system.'}, status=404)
                
                # Check if already verified
                if email_address.verified:
                    return JsonResponse({'success': True, 'message': 'Email already verified.'})

                email_address.send_confirmation(request)
                return JsonResponse({'success': True, 'message': 'Verification email sent!'})

            except Exception as e:
                # Log the actual error for debugging
                
                print(traceback.format_exc())
                return JsonResponse({'success': False, 'error': f'Send failed: {str(e)}'}, status=500)
                
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Server Error: {str(e)}'}, status=500)
            
    return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=400)

def demo_login(request):
    """
    Logs in the read-only 'demo' user without password authentication.
    Ensures data is always fresh (current month).
    """
    # Clear messages
    list(messages.get_messages(request))

    try:
        user = User.objects.get(username='demo')
        
        # Check if data is stale (i.e. not from this month)
        # We check the latest expense. If no expenses or old date, refresh.
        last_expense = Expense.objects.filter(user=user).order_by('-date').first()
        is_stale = False
        
        if not last_expense:
            is_stale = True
        else:
            today = date.today()
            if last_expense.date.month != today.month or last_expense.date.year != today.year:
                is_stale = True
        
        if is_stale:
            # Data is old, refresh it
            call_command('setup_demo_user')
            # Refetch the new user object since the old one might have been deleted/recreated
            user = User.objects.get(username='demo')

    except User.DoesNotExist:
        # User doesn't exist, create it
        call_command('setup_demo_user')
        user = User.objects.get(username='demo')

    # Manually set the backend to allow login without authentication
    login(request, user, backend='django.contrib.auth.backends.ModelBackend')
    messages.success(request, "🚀 Welcome to Demo Mode! Feel free to explore the app.")
    return redirect('home')

def demo_signup(request):
    """
    Logs out the demo user and redirects to the signup page.
    """
    logout(request)
    return redirect('signup')

# --------------------
# Mixins
# --------------------

class RecurringTransactionMixin:
    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            self.process_recurring_transactions(request.user)
        return super().dispatch(request, *args, **kwargs)

    def process_recurring_transactions(self, user):
        today = date.today()
        recurring_txs = RecurringTransaction.objects.filter(user=user, is_active=True)
        
        for rt in recurring_txs:
            if not rt.last_processed_date:
                current_date = rt.start_date
            else:
                current_date = rt.get_next_date(rt.last_processed_date, rt.frequency)

            while current_date <= today:
                description = f"{rt.description} (Recurring)"
                if rt.transaction_type == 'EXPENSE':
                    Expense.objects.get_or_create(
                        user=user,
                        date=current_date,
                        amount=rt.amount,
                        category=rt.category or 'Uncategorized',
                        defaults={
                            'description': description,
                            'payment_method': rt.payment_method
                        }
                    )
                else:
                    Income.objects.get_or_create(
                        user=user,
                        date=current_date,
                        amount=rt.amount,
                        source=rt.source or 'Other',
                        defaults={'description': description}
                    )
                
                rt.last_processed_date = current_date
                rt.save()
                current_date = rt.get_next_date(current_date, rt.frequency)


# Custom signup view to log user in immediately
class SignUpView(generic.CreateView):
    form_class = CustomSignupForm
    success_url = reverse_lazy('account_login')
    template_name = 'registration/signup.html'

class LandingPageView(TemplateView):
    template_name = 'landing.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            return redirect('home')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        plans = SubscriptionPlan.objects.filter(is_active=True)
        context['plans'] = {p.tier: p for p in plans}
        return context

class SettingsHomeView(LoginRequiredMixin, TemplateView):
    template_name = 'expenses/settings_home.html'

@login_required
def home_view(request):
    """
    Dashboard view with filters and multiple charts.
    """
    from decimal import Decimal

    # Base QuerySet
    expenses = Expense.objects.filter(user=request.user).order_by('-date')
    
    # Filter Logic
    selected_years = request.GET.getlist('year')
    selected_months = request.GET.getlist('month')
    selected_categories = request.GET.getlist('category')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Remove empty strings from lists
    selected_years = [y for y in selected_years if y]
    selected_months = [m for m in selected_months if m]
    selected_categories = [c for c in selected_categories if c]

    # Date Range takes precedence
    if start_date or end_date:
        if start_date:
            expenses = expenses.filter(date__gte=start_date)
        if end_date:
            expenses = expenses.filter(date__lte=end_date)
        
        # Reset lists for UI clarity since we are in range mode
        selected_years = []
        selected_months = []
        
        trend_title = "Expenses Trend (Custom Range)"
    else:
        # Default to current month/year ONLY on initial land (no params)
        if not request.GET and not (selected_years or selected_months):
            selected_years = [str(datetime.now().year)]
            selected_months = [str(datetime.now().month)]
        
        if selected_years:
            expenses = expenses.filter(date__year__in=selected_years)
        if selected_months:
            expenses = expenses.filter(date__month__in=selected_months)
            
        if len(selected_months) == 1 and len(selected_years) == 1:
            trend_title = f"Daily Expenses for {selected_months[0]}/{selected_years[0]}"
        else:
            trend_title = "Monthly Expenses Trend"

    if selected_categories:
        expenses = expenses.filter(category__in=selected_categories)
        
    # Income Logic (Mirroring Expense Filters)
    incomes = Income.objects.filter(user=request.user)
    if start_date or end_date:
        if start_date:
            incomes = incomes.filter(date__gte=start_date)
        if end_date:
            incomes = incomes.filter(date__lte=end_date)
    else:
        if selected_years:
            incomes = incomes.filter(date__year__in=selected_years)
        if selected_months:
            incomes = incomes.filter(date__month__in=selected_months)
    
    total_income = incomes.aggregate(Sum('amount'))['amount__sum'] or 0
    all_dates = Expense.objects.filter(user=request.user).dates('date', 'year', order='DESC')
    years = sorted(list(set([d.year for d in all_dates] + [datetime.now().year])), reverse=True)
    all_categories = Expense.objects.filter(user=request.user).values_list('category', flat=True).distinct().order_by('category')

    # 1. Category Chart Data (Distribution) & Summary Table
    # We need to fetch raw values and merge them in Python to handle whitespace duplicates
    raw_category_data = expenses.values('category').annotate(total=Sum('amount'))
    
    # Process and merge duplicates
    merged_category_map = {}
    for item in raw_category_data:
        # Strip whitespace to normalize
        cat_name = item['category'].strip()
        amount = float(item['total'])
        
        if cat_name in merged_category_map:
            merged_category_map[cat_name] += amount
        else:
            merged_category_map[cat_name] = amount
            
    # Convert back to list of dicts for template/charts, sorted by total
    # This replaces the DB-ordered queryset with a sorted list
    category_data = [
        {'category': cat, 'total': amt} 
        for cat, amt in merged_category_map.items()
    ]
    category_data.sort(key=lambda x: x['total'], reverse=True)

    # Compute limits and usage per category for chart display
    # Compute limits and usage per category for chart display
    category_limits = []
    # Optimization: Pre-fetch all categories for the user to avoid N+1 queries in the loop
    user_categories = {c.name: c for c in Category.objects.filter(user=request.user)}

    for item in category_data:
        cat_name = item['category']
        cat_obj = user_categories.get(cat_name)

        limit = float(cat_obj.limit) if (cat_obj and cat_obj.limit) else None

        used_percent = round((item['total'] / limit * 100), 1) if limit else None
        category_limits.append({
            'name': cat_name,
            'total': item['total'],
            'limit': limit,
            'used_percent': used_percent,
        })
    
    categories = [item['category'] for item in category_data]
    category_amounts = [item['total'] for item in category_data]
    
    # 2. Time Trend (Stacked) Data
    
    # Determine Labels (X-Axis)
    # Determine Labels (X-Axis)
    if start_date or end_date:
        # For custom range, if range < 60 days, show daily. Else monthly.
        # Simple heuristic: Always show daily for custom range for now, or let logic decide.
        # Let's stick to: if explicit month selected -> daily. If range -> daily (usually granular).
        trend_qs = expenses.annotate(period=TruncDay('date'))
        date_format = '%d %b'
    elif len(selected_months) == 1 and len(selected_years) == 1:
        # Daily view
        trend_qs = expenses.annotate(period=TruncDay('date'))
        date_format = '%d %b'
    else:
        # Monthly view
        trend_qs = expenses.annotate(period=TruncMonth('date'))
        date_format = '%b %Y'

    # Aggregate by Period AND Category for Stacking
    stacked_data = trend_qs.values('period', 'category').annotate(total=Sum('amount')).order_by('period')
    
    # Process into Chart.js Datasets
    # 1. Get unique sorted periods
    periods = sorted(list(set(item['period'] for item in stacked_data)))
    trend_labels = [p.strftime(date_format) for p in periods]
    
    # 2. Build datasets map: { 'CategoryA': [0, 10, 0...], 'CategoryB': ... }
    # Initialize with zeros for all unique NORMALIZED categories found in expenses
    normalized_all_categories = sorted(list(merged_category_map.keys()))
    dataset_map = { cat: [0] * len(periods) for cat in normalized_all_categories }
    
    for item in stacked_data:
        p_idx = periods.index(item['period'])
        # Strip to match our normalized keys
        cat = item['category'].strip()
        if cat in dataset_map:
            dataset_map[cat][p_idx] += float(item['total']) # Add += in case multiple unstripped cats map to same striped cat in same period
            
    # 3. Convert map to list of dataset objects for Chart.js
    trend_datasets = []
    # Define a color palette (Light Blue, Blue Green, Prussian Blue, Honey Yellow, Orange)
    colors = ['#219EBC', '#023047', '#8ECAE6', '#FFB703', '#0575E6']
    
    for i, (cat, data) in enumerate(dataset_map.items()):
        # Only include non-zero datasets
        if sum(data) > 0:
             trend_datasets.append({
                 'label': cat,
                 'data': data,
                 'backgroundColor': colors[i % len(colors)],
                 'borderRadius': 2
             })

    # 3. Top 5 Expenses
    top_expenses_qs = expenses.order_by('-amount')[:5]
    top_labels = [e.description[:20] + '...' if len(e.description) > 20 else e.description for e in top_expenses_qs]
    top_amounts = [float(e.amount) for e in top_expenses_qs]

    # --- NEW: Income vs Expenses Trend Data ---
    # Re-use the truncation logic determined above
    if start_date or end_date or (len(selected_months) == 1 and len(selected_years) == 1):
        trunc_func = TruncDay
    else:
        trunc_func = TruncMonth
        
    inc_trend = incomes.annotate(period=trunc_func('date')).values('period').annotate(total=Sum('amount')).order_by('period')
    exp_trend = expenses.annotate(period=trunc_func('date')).values('period').annotate(total=Sum('amount')).order_by('period')
    
    # Merge periods
    inc_periods = set(i['period'] for i in inc_trend)
    exp_periods = set(e['period'] for e in exp_trend)
    all_periods_sorted = sorted(list(inc_periods.union(exp_periods)))
    
    ie_labels = [p.strftime(date_format) for p in all_periods_sorted]
    ie_income_data = [float(inc_trend.get(period=p)['total']) if inc_trend.filter(period=p).exists() else 0 for p in all_periods_sorted]
    # Optimization: Use dict lookup instead of filter inside loop
    inc_map = {i['period']: float(i['total']) for i in inc_trend}
    exp_map = {e['period']: float(e['total']) for e in exp_trend}
    
    ie_income_data = [inc_map.get(p, 0.0) for p in all_periods_sorted]
    ie_expense_data = [exp_map.get(p, 0.0) for p in all_periods_sorted]
    ie_savings_data = [inc_map.get(p, 0.0) - exp_map.get(p, 0.0) for p in all_periods_sorted]

    # --- 1. Payment Method Distribution (method -> sources; use only payment_source) ---
    # Cash when payment_method is Cash (payment_source can be null). Others use payment_source to resolve card or bank.
    expenses_qs = expenses.values('id', 'payment_method', 'amount', 'payment_source')
    payment_map = {}  # method -> { source_name: total }
    for expense in expenses_qs:
        pm_name = expense['payment_method'] or 'Unknown'
        amount = float(expense['amount'])
        source_id = expense.get('payment_source')

        if pm_name == 'Cash':
            # Cash: no payment_source needed; if method is Cash it's cash
            if pm_name not in payment_map:
                payment_map[pm_name] = {}
            payment_map[pm_name]['Cash'] = payment_map[pm_name].get('Cash', 0) + amount
        elif pm_name == 'Credit Card':
            if source_id:
                card = CreditCard.objects.filter(user=request.user, pk=source_id).first()
                if card:
                    if pm_name not in payment_map:
                        payment_map[pm_name] = {}
                    payment_map[pm_name][card.name] = payment_map[pm_name].get(card.name, 0) + amount
                else:
                    if pm_name not in payment_map:
                        payment_map[pm_name] = {}
                    payment_map[pm_name]['Unknown'] = payment_map[pm_name].get('Unknown', 0) + amount
            else:
                if pm_name not in payment_map:
                    payment_map[pm_name] = {}
                payment_map[pm_name]['Unknown'] = payment_map[pm_name].get('Unknown', 0) + amount
        else:
            # UPI, Debit Card, NetBanking: payment_source = PaymentSource id
            if pm_name not in payment_map:
                payment_map[pm_name] = {}
            if source_id:
                source = PaymentSource.objects.filter(user=request.user, pk=source_id).first()
                if source:
                    payment_map[pm_name][source.name] = payment_map[pm_name].get(source.name, 0) + amount
                else:
                    payment_map[pm_name]['Unknown'] = payment_map[pm_name].get('Unknown', 0) + amount
            else:
                payment_map[pm_name]['Unknown'] = payment_map[pm_name].get('Unknown', 0) + amount

    # Stacked bar: one row per payment method, segments = sources (cards/banks)
    method_totals = [(m, sum(sub.values())) for m, sub in payment_map.items()]
    method_totals.sort(key=lambda x: x[1], reverse=True)
    payment_method_labels = [m for m, _ in method_totals]
    method_index = {m: i for i, m in enumerate(payment_method_labels)}
    n_methods = len(payment_method_labels)
    payment_stacked_datasets = []
    for method, segments in payment_map.items():
        idx = method_index[method]
        for segment_name, amount in segments.items():
            data = [0.0] * n_methods
            data[idx] = amount
            payment_stacked_datasets.append({"label": segment_name, "data": data})
    payment_flat = []
    for method, segments in payment_map.items():
        for seg_name, amt in segments.items():
            payment_flat.append((f"{method} - {seg_name}", amt))
    sorted_payment_items = sorted(payment_flat, key=lambda x: x[1], reverse=True)
    payment_labels = [item[0] for item in sorted_payment_items]
    payment_data = [item[1] for item in sorted_payment_items]

    # --- 2. Cashback by Payment Source (which card/bank has more cashback) ---
    cashback_qs = expenses.filter(has_cashback=True).values(
        'payment_method', 'payment_source', 'amount', 'cashback_type', 'cashback_value'
    )
    cashback_by_source = {}  # source_name -> total_cashback
    for row in cashback_qs:
        if not row.get('cashback_value'):
            continue
        amt = Decimal(str(row['amount']))
        cb_type = row.get('cashback_type')
        cb_val = Decimal(str(row['cashback_value']))
        if cb_type == 'PERCENTAGE':
            cashback_amt = float((amt * cb_val) / 100)
        elif cb_type == 'FIXED':
            cashback_amt = float(cb_val)
        else:
            continue
        pm_name = row.get('payment_method') or 'Unknown'
        source_id = row.get('payment_source')
        if pm_name == 'Credit Card' and source_id:
            card = CreditCard.objects.filter(user=request.user, pk=source_id).first()
            name = card.name if card else 'Unknown'
        elif source_id:
            source = PaymentSource.objects.filter(user=request.user, pk=source_id).first()
            name = source.name if source else 'Unknown'
        else:
            name = 'Cash/Other'
        cashback_by_source[name] = cashback_by_source.get(name, 0) + cashback_amt
    cashback_source_labels = list(cashback_by_source.keys())
    cashback_source_data = [cashback_by_source[k] for k in cashback_source_labels]
    # Sort by cashback amount desc
    if cashback_source_labels:
        combined = sorted(zip(cashback_source_labels, cashback_source_data), key=lambda x: x[1], reverse=True)
        cashback_source_labels = [x[0] for x in combined]
        cashback_source_data = [x[1] for x in combined]


    # 4. Summary Stats
    total_expenses = expenses.aggregate(Sum('amount'))['amount__sum'] or 0
    transaction_count = expenses.count()
    top_category = category_data[0] if category_data else None
    
    savings = total_income - total_expenses

    # --- NEW: Savings Projection (Linear Extrapolation) ---
    current_date = date.today()
    current_year = current_date.year
    current_month = current_date.month 

    # 1. Calculate YTD Savings (Strictly for current year, regardless of filters)
    ytd_income = Income.objects.filter(user=request.user, date__year=current_year, date__month__lte=current_month).aggregate(Sum('amount'))['amount__sum'] or 0
    ytd_expenses = Expense.objects.filter(user=request.user, date__year=current_year, date__month__lte=current_month).aggregate(Sum('amount'))['amount__sum'] or 0
    ytd_savings = ytd_income - ytd_expenses
    
    projected_savings = 0
    
    # Only project if we have data and positive savings
    if ytd_savings > 0:
        # Avoid division by zero if it's January (month 1)
        # Actually, even in Jan, months_passed is 1. So we are good.
        months_passed = current_month
        avg_monthly_savings = ytd_savings / months_passed
        
        months_remaining = 12 - months_passed
        projected_additional = avg_monthly_savings * months_remaining
        
        projected_savings = ytd_savings + projected_additional
    else:
        # If savings are negative or zero, projection is effectively "0" or "current state"
        # We might handle this in template
        projected_savings = 0

    # Calculate MoM Changes ONLY if exactly one year and one month are selected
    prev_month_data = None
    if len(selected_years) == 1 and len(selected_months) == 1:
        try:
            sel_year = int(selected_years[0])
            sel_month = int(selected_months[0])
            
            # Calculate previous month and year
            if sel_month == 1:
                prev_month = 12
                prev_year = sel_year - 1
            else:
                prev_month = sel_month - 1
                prev_year = sel_year

            prev_expenses = Expense.objects.filter(user=request.user, date__year=prev_year, date__month=prev_month).aggregate(Sum('amount'))['amount__sum'] or 0
            prev_income = Income.objects.filter(user=request.user, date__year=prev_year, date__month=prev_month).aggregate(Sum('amount'))['amount__sum'] or 0
            prev_savings = prev_income - prev_expenses

            def calc_pct(current, previous):
                if previous == 0:
                    return None
                return ((current - previous) / previous) * 100

            prev_month_data = {
                'income_pct': calc_pct(total_income, prev_income),
                'expense_pct': calc_pct(total_expenses, prev_expenses),
                'savings_pct': calc_pct(savings, prev_savings),
            }
            # Add absolute values for template display
            for key in list(prev_month_data.keys()):
                val = prev_month_data[key]
                if val is not None:
                    prev_month_data[f'{key}_abs'] = abs(val)
        except (ValueError, IndexError):
            pass

    # Prepare display labels for the template
    display_year = None
    display_month = None
    
    if len(selected_years) == 1:
        display_year = selected_years[0]
        
    if len(selected_months) == 1:
        try:
            m_idx = int(selected_months[0])
            display_month = calendar.month_name[m_idx]
        except (ValueError, IndexError):
            pass

    # NEW: Calculate Previous/Next Month URLs
    prev_month_url = None
    next_month_url = None

    if len(selected_years) == 1 and len(selected_months) == 1:
        try:
            curr_year = int(selected_years[0])
            curr_month = int(selected_months[0])
            
            # Previous Month
            if curr_month == 1:
                pm = 12
                py = curr_year - 1
            else:
                pm = curr_month - 1
                py = curr_year
            
            # Next Month
            if curr_month == 12:
                nm = 1
                ny = curr_year + 1
            else:
                nm = curr_month + 1
                ny = curr_year

            # Construct Query String (Preserve Categories)
            base_qs = []
            for c in selected_categories:
                base_qs.append(f'category={c}')
            
            qs_prev = base_qs + [f'year={py}', f'month={pm}']
            qs_next = base_qs + [f'year={ny}', f'month={nm}']
            
            prev_month_url = f"{reverse('home')}?{'&'.join(qs_prev)}"
            next_month_url = f"{reverse('home')}?{'&'.join(qs_next)}"
            
        except ValueError:
            pass
    
    # --- Emotional Feedback / Insights Logic (Enhanced) ---
    
    insights = []
    
    # helper for streaks
    def get_monthly_savings_status(u, y, m):
        inc = Income.objects.filter(user=u, date__year=y, date__month=m).aggregate(Sum('amount'))['amount__sum'] or 0
        exp = Expense.objects.filter(user=u, date__year=y, date__month=m).aggregate(Sum('amount'))['amount__sum'] or 0
        return inc > exp

    # Construct date params for deep linking
    date_params = ""
    for y in selected_years:
        date_params += f"&year={y}"
    for m in selected_months:
        date_params += f"&month={m}"

    # helper for category links
    def link_cats(cats):
        links_html = format_html_join(
            mark_safe(', '),
            '<a href="{}" class="alert-link text-decoration-underline">{}</a>',
            ((reverse('expense-list') + f"?category={c}{date_params}", c) for c in cats[:2])
        )
        if len(cats) > 2:
            return format_html('{}, etc.', links_html)
        return links_html

    # 0. Anomaly Detection (Spending Spike)
    # Only if viewing current month (or default view)
    is_current_month_view = False
    now = datetime.now()
    if not request.GET or (len(selected_months) == 1 and str(now.month) in selected_months and str(now.year) in selected_years):
         is_current_month_view = True
    
    if is_current_month_view and total_expenses > 0:
        # Calculate last 3 months average
        last_3_months_total = 0
        months_counted = 0
        for i in range(1, 4):
            # Calculate past month/year
            y = now.year
            m = now.month - i
            while m < 1:
                m += 12
                y -= 1
            
            m_total = Expense.objects.filter(user=request.user, date__year=y, date__month=m).aggregate(Sum('amount'))['amount__sum'] or 0
            if m_total > 0:
                last_3_months_total += m_total
                months_counted += 1
        
        if months_counted > 0:
            avg_past_spend = last_3_months_total / months_counted
            
            # Project current month
            days_in_month = calendar.monthrange(now.year, now.month)[1]
            days_passed = now.day
            if days_passed > 0:
                projected_spend = (float(total_expenses) / days_passed) * days_in_month
                avg_past_spend_float = float(avg_past_spend)
                
                if projected_spend > avg_past_spend_float * 1.25 and float(total_expenses) > 1000: # 25% Higher + Min Threshold
                    pct_higher = int(((projected_spend - avg_past_spend_float) / avg_past_spend_float) * 100)
                    insights.append({
                        'type': 'warning',
                        'icon': 'graph-up-arrow',
                        'title': 'Traffic Alert 🚦',
                        'message': f"You're pacing {pct_higher}% higher than usual. Slow down to stay on track!",
                        'allow_share': False
                    })

    # 1. Budget Warnings (High Priority)

    over_budget_cats = [c['name'] for c in category_limits if c['used_percent'] is not None and c['used_percent'] > 100]
    near_budget_cats = [c['name'] for c in category_limits if c['used_percent'] is not None and 90 <= c['used_percent'] <= 100]
    
    # Check savings rate for "Softener" context
    savings_rate = (savings / total_income * 100) if total_income > 0 else 0
    
    if over_budget_cats:
        cats_str = link_cats(over_budget_cats)
        
        if savings_rate >= 20:
            # Contextualized Warning for High Savers
            msg = format_html("Even strong months have leaks. You crossed limits in {} — catching this keeps you on track.", cats_str)
        else:
            # Standard Coaching Warning - "Warning" type (Yellow) instead of Danger (Red) for empathy
            msg = format_html("⚠️ Budget crossed in {} — let’s rebalance to stay safe.", cats_str)

        insights.append({
            'type': 'warning', # Changed from danger
            'icon': 'exclamation-octagon-fill',
            'title': 'Budget Breached',
            'message': msg,
            'allow_share': False
        })
    elif near_budget_cats:
        cats_str = link_cats(near_budget_cats)
        insights.append({
            'type': 'warning',
            'icon': 'exclamation-triangle-fill',
            'title': 'Approaching Limit',
            'message': format_html("Heads up! You're close to overspending on {}.", cats_str),
            'allow_share': False
        })

    # 2. Wins & Cause-Based Praise (Specific & Celebratory)
    if prev_month_data:
        # Calculate Category Savings (Cause of the win)
        # We need prev month category breakdown
        prev_cat_qs = Expense.objects.filter(user=request.user, date__year=prev_year, date__month=prev_month).values('category').annotate(total=Sum('amount'))
        prev_cat_map = {item['category'].strip(): float(item['total']) for item in prev_cat_qs}
        
        savings_contributors = []
        for cat, curr_total in merged_category_map.items():
            prev_total = prev_cat_map.get(cat, 0)
            if prev_total > curr_total:
                diff = prev_total - curr_total
                if diff > 100: # Threshold to mention
                    savings_contributors.append((cat, diff))
        savings_contributors.sort(key=lambda x: x[1], reverse=True)
        top_savers = [c[0] for c in savings_contributors[:2]]
        
        # Savings Win
        if total_income > 0 and savings > 0:
            savings_rate = (savings / total_income) * 100
            if savings_rate >= 20:
                msg_text = f"You've saved {savings_rate:.0f}% of your income this month."
                share_text = f"I saved {savings_rate:.0f}% of my income this month using TrackMyRupee! 🏆"
                
                if top_savers:
                    cats_link = link_cats(top_savers)
                    msg = format_html("{} You spent less on {} — that's where the magic happened.", msg_text, cats_link)
                else:
                    msg = msg_text

                insights.append({
                    'type': 'success',
                    'icon': 'trophy-fill',
                    'title': 'Super Saver Status! 🏆',
                    'message': msg,
                    'allow_share': True,
                    'share_text': share_text
                })
            elif prev_month_data['savings_pct'] and prev_month_data['savings_pct'] > 0:
                 insights.append({
                    'type': 'success',
                    'icon': 'graph-up-arrow',
                    'title': 'Momentum Building 🚀',
                    'message': f"Your savings grew by {prev_month_data['savings_pct_abs']:.0f}% vs last month. You're getting better at this!",
                    'allow_share': True,
                    'share_text': f"My savings grew by {prev_month_data['savings_pct_abs']:.0f}% this month! 🚀 via TrackMyRupee"
                })
        
        # Expense Control Win (if we haven't already praised savings)
        if len(insights) == 0: 
            if prev_month_data['expense_pct'] and prev_month_data['expense_pct'] < -5:
                 msg_text = f"You've cut spending by {prev_month_data['expense_pct_abs']:.0f}%."
                 share_text = f"I cut my spending by {prev_month_data['expense_pct_abs']:.0f}% this month! 👍 via TrackMyRupee"
                 
                 if top_savers:
                     cats_link = link_cats(top_savers)
                     msg = format_html("{} {} saw the biggest drops.", msg_text, cats_link)
                 else:
                     msg = msg_text
                 
                 insights.append({
                    'type': 'success',
                    'icon': 'check-circle-fill',
                    'title': 'You’re in Control 👍',
                    'message': msg,
                    'allow_share': True,
                    'share_text': share_text
                })

    # 3. Streak & Identity (Reassuring / Habit Forming)
    # Only calculate if current status is good
    if savings > 0 and len(selected_years) == 1 and len(selected_months) == 1:
        streak = 1 # Current month counts
        check_to_go = 5 # check max 5 months back
        curr_y_calc, curr_m_calc = int(selected_years[0]), int(selected_months[0])
        
        for i in range(check_to_go):
            # Go back one month
            if curr_m_calc == 1:
                curr_m_calc = 12
                curr_y_calc -= 1
            else:
                curr_m_calc -= 1
            
            if get_monthly_savings_status(request.user, curr_y_calc, curr_m_calc):
                streak += 1
            else:
                break
        
        if streak > 1:
            insights.append({
                'type': 'info', # Use Info for "Identity/Streak"
                'icon': 'fire',
                'title': 'On a Roll!',
                'message': f"🔥 This is your {streak}th month in a row staying under budget.",
                'allow_share': True,
                'share_text': f"🔥 I've stayed under budget for {streak} months in a row! via TrackMyRupee"
            })

    # 4. Fallback
    if not insights and savings > 0:
        insights.append({
            'type': 'info',
            'icon': 'piggy-bank-fill',
            'title': 'In the Green',
            'message': f"You've saved {savings} so far. Keep it up!",
            'allow_share': False
        })
    elif not insights:
        insights.append({
            'type': 'secondary',
            'icon': 'stars',
            'title': 'Fresh Start',
            'message': "Small steps today lead to big results tomorrow. Let's track some expenses!",
            'allow_share': False
        })

    # Limit to top 2 insights to avoid clutter
    insights = insights[:2]

    # Check for onboarding (True if user has NO data at all)
    has_any_data = Expense.objects.filter(user=request.user).exists() or Income.objects.filter(user=request.user).exists()

    context = {
        'is_new_user': not has_any_data,
        'insights': insights[::-1],
        'total_income': total_income,
        'savings': savings,
        'recent_transactions': expenses.order_by('-date')[:5],
        'categories': categories,
        'category_amounts': category_amounts,
        'category_data': category_data, # Passing full queryset for the summary table
        'category_limits': category_limits,
        'trend_labels': trend_labels,
        'trend_datasets': trend_datasets,
        'trend_title': trend_title,
        'top_labels': top_labels,
        'top_amounts': top_amounts,
        # New Context
        'ie_labels': ie_labels,
        'ie_income_data': ie_income_data,
        'ie_expense_data': ie_expense_data,
        'ie_savings_data': ie_savings_data,
        'payment_labels': payment_labels,
        'payment_data': payment_data,
        'payment_method_labels': payment_method_labels,
        'payment_stacked_datasets': payment_stacked_datasets,
        'cashback_source_labels': cashback_source_labels,
        'cashback_source_data': cashback_source_data,
        'years': years,
        'all_categories': all_categories,
        'selected_years': selected_years,
        'selected_months': selected_months,
        'selected_year': display_year,    # NEW: For template display labels
        'selected_month': display_month,  # NEW: For template display labels
        'selected_categories': selected_categories,
        'months_list': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'total_expenses': total_expenses,
        'transaction_count': transaction_count,
        'top_category': top_category,
        'projected_savings': projected_savings, # NEW
        'start_date': start_date,
        'end_date': end_date,
        'prev_month_data': prev_month_data,
        'prev_month_url': prev_month_url,
        'next_month_url': next_month_url,
        'show_tutorial': not request.user.profile.has_seen_tutorial or request.GET.get('tour') == 'true',
        'has_any_budget': any((c.get('limit') or 0) > 0 for c in category_limits),
    }
    return render(request, 'home.html', context)

@login_required
def complete_tutorial(request):
    if request.method == 'POST':
        profile, created = UserProfile.objects.get_or_create(user=request.user)
        profile.has_seen_tutorial = True
        profile.save()
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'}, status=400)

@login_required
def upload_view(request):
    """
    Upload view with year selection enforcement.
    """
    
    if request.method == 'POST' and request.FILES.get('file'):
        excel_file = request.FILES['file']
        selected_year = int(request.POST.get('year'))
        
        try:
            # Load workbook
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = list(sheet.iter_rows(values_only=True))
                
                if not rows:
                    continue

                # Search for the header row index
                header_row_index = -1
                header_cols = []
                
                for i, row in enumerate(rows[:10]):
                    if not row: continue
                    row_values = [str(val).strip().title() if val is not None else "" for val in row]
                    if 'Date' in row_values and 'Amount' in row_values and 'Description' in row_values:
                        header_row_index = i
                        header_cols = row_values
                        break
                
                if header_row_index == -1:
                    print(f"Skipping sheet {sheet_name}: Could not find header row.")
                    continue

                # Map column indices
                col_map = {col: idx for idx, col in enumerate(header_cols) if col}
                required_columns = ['Date', 'Amount', 'Description', 'Category']
                
                if not all(col in col_map for col in required_columns):
                    print(f"Skipping sheet {sheet_name}: Missing required columns.")
                    continue

                # Process data rows
                for row_data in rows[header_row_index + 1:]:
                    if not any(row_data): continue # Skip empty rows
                    
                    # Parse date
                    date_val = row_data[col_map['Date']]
                    if date_val is None:
                        continue
                        
                    date_obj = None
                    if isinstance(date_val, str):
                        formats = ['%d %b %Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y', '%d %B %Y', '%d %b', '%d-%b', '%d %B']
                        for fmt in formats:
                            try:
                                parsed_date = datetime.strptime(date_val.strip(), fmt).date()
                                date_obj = parsed_date.replace(year=selected_year)
                                break
                            except ValueError:
                                continue
                        if not date_obj:
                            continue
                    elif isinstance(date_val, (datetime, date)):
                        date_obj = date_val if isinstance(date_val, date) else date_val.date()
                        try:
                            date_obj = date_obj.replace(year=selected_year)
                        except ValueError:
                            date_obj = date_obj.replace(day=28, year=selected_year)
                    else:
                        continue # Unsupported date type

                    # Get other fields
                    amount = row_data[col_map['Amount']]
                    description = row_data[col_map['Description']]
                    category = row_data[col_map['Category']] if 'Category' in col_map else None
                    
                    if amount is None or description is None:
                        continue

                    category_obj = None
                    if category:
                        category_name = str(category).strip()
                        if category_name:
                            category_obj, _ = Category.objects.get_or_create(user=request.user, name=category_name)

                    Expense.objects.get_or_create(
                        user=request.user,
                        date=date_obj,
                        amount=float(amount) if not isinstance(amount, float) else amount,
                        description=str(description),
                        category=category_obj.name if category_obj else "Others"
                    )
            return redirect('home')
        except Exception as e:
            print(f"Error processing file: {e}")
            traceback.print_exc()
            pass

    # Context for year dropdown
    current_year = datetime.now().year
    years = range(current_year, current_year - 5, -1)
    
    return render(request, 'upload.html', {'years': years, 'current_year': current_year})

class ExpenseListView(LoginRequiredMixin, RecurringTransactionMixin, ListView):
    model = Expense
    template_name = "expenses/expense_list.html"
    context_object_name = "expenses"
    paginate_by = 20

    def get_queryset(self):
        queryset = (
            Expense.objects.filter(user=self.request.user)
            .select_related("shared_details")
            .order_by("-date")
        )
        
        # Filtering
        selected_years = self.request.GET.getlist('year')
        selected_months = self.request.GET.getlist('month')
        selected_categories = self.request.GET.getlist('category')
        search_query = self.request.GET.get('search')
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')

        # Remove empty strings from lists
        selected_years = [y for y in selected_years if y]
        selected_months = [m for m in selected_months if m]
        selected_categories = [c for c in selected_categories if c]
        
        # Date Range Logic (Precedence over Year/Month)
        if start_date or end_date:
            if start_date:
                queryset = queryset.filter(date__gte=start_date)
            if end_date:
                queryset = queryset.filter(date__lte=end_date)
        else:
            # Check if any specific filter is active
            has_active_filters = (
                selected_years or 
                selected_months or 
                search_query  # Don't check categories as we might want defaults even if cat is selected? No, usually filters are additive.
            )
            
            # If no year/month/search filters, default to current month/year
            # (ignoring category here might be debated, but typically if I just filter 'Food', I might want all time or current month? 
            #  The dashboard logic defaults to current month if no year/month. Let's stick to that.)
            if not has_active_filters:
                selected_years = [str(datetime.now().year)]
                selected_months = [str(datetime.now().month)]
            
            if selected_years:
                queryset = queryset.filter(date__year__in=selected_years)
            
            if selected_months:
                queryset = queryset.filter(date__month__in=selected_months)

        if selected_categories:
            queryset = queryset.filter(category__in=selected_categories)
        
        # Filter by Payment Method
        payment_method = self.request.GET.get('payment_method')
        if payment_method:
            queryset = queryset.filter(payment_method=payment_method)

        if search_query:
            queryset = queryset.filter(description__icontains=search_query)
            
        # Sorting
        sort_by = self.request.GET.get('sort')
        if sort_by == 'amount_asc':
            queryset = queryset.order_by('amount')
        elif sort_by == 'amount_desc':
            queryset = queryset.order_by('-amount')
        # Default is already '-date' from line 961, so valid fallback.
            
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Calculate stats for the filtered queryset
        filtered_queryset = self.object_list
        context['filtered_count'] = filtered_queryset.count()
        context['filtered_amount'] = filtered_queryset.aggregate(Sum('amount'))['amount__sum'] or 0

        # Get unique years and categories for validation
        user_expenses = Expense.objects.filter(user=self.request.user)
        years_dates = user_expenses.dates('date', 'year', order='DESC')
        years = sorted(list(set([d.year for d in years_dates] + [datetime.now().year])), reverse=True)
        # Python-side deduplication to handle whitespace variants (e.g. "Goa" vs "Goa ")
        raw_categories = user_expenses.values_list('category', flat=True)
        categories = sorted(list(set([c.strip() for c in raw_categories if c and c.strip()])), key=str.lower)
        
        context['years'] = years
        context['categories'] = categories
        context['months_list'] = [(i, calendar.month_name[i]) for i in range(1, 13)]
        
        # Determine selected year for UI
        # Determine selected year for UI
        year_param = self.request.GET.get('year')
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')

        context['start_date'] = start_date
        context['end_date'] = end_date
        
        if start_date or end_date:
            context['selected_years'] = []
            context['selected_months'] = []
            context['selected_categories'] = []
        else:
            selected_years = self.request.GET.getlist('year')
            selected_months = self.request.GET.getlist('month')
            selected_categories = self.request.GET.getlist('category')
            search_query = self.request.GET.get('search')
            
            # Remove empty strings
            selected_years = [y for y in selected_years if y]
            selected_months = [m for m in selected_months if m]
            selected_categories = [c for c in selected_categories if c]

            # Check if any specific filter is active
            has_active_filters = (
                selected_years or 
                selected_months or 
                search_query 
                # (ignoring category here as well to match get_queryset)
            )

            # Mirror default logic from get_queryset
            if not has_active_filters:
                selected_years = [str(datetime.now().year)]
                selected_months = [str(datetime.now().month)]
            
            context['selected_years'] = selected_years
            context['selected_months'] = selected_months
            context['selected_categories'] = selected_categories
            
        return context

class ExpenseCreateView(LoginRequiredMixin, generic.TemplateView):
    template_name = 'expenses/expense_form.html'

    def get(self, request, *args, **kwargs):
        # Check if bulk mode is requested (for backward compatibility)
        bulk_mode = request.GET.get("bulk", "false").lower() == "true"

        if bulk_mode:
            # Use formset for bulk entry
            ExpenseFormSet = modelformset_factory(
                Expense, form=ExpenseForm, extra=1, can_delete=True
            )
            initial_data = [{"date": datetime.now().date()} for _ in range(1)]
            formset = ExpenseFormSet(
                queryset=Expense.objects.none(),
                initial=initial_data,
                form_kwargs={"user": request.user},
            )
            next_url = request.GET.get("next", "")
            return render(
                request, self.template_name, {"formset": formset, "next_url": next_url}
            )
        else:
            # Check if this is a copy operation
            copy_expense_id = request.GET.get("copy")

            # Prepare initial data if copying
            initial_data = {}
            copy_shared_data = None

            if copy_expense_id:
                try:
                    # Fetch the expense to copy
                    expense = get_object_or_404(
                        Expense, pk=copy_expense_id, user=request.user
                    )

                    initial_data = {
                        "date": expense.date,
                        "amount": expense.amount,
                        "description": f"{expense.description} (Copy)",
                        "category": expense.category,
                        "payment_method": expense.payment_method,
                    }

                    # Add cashback data if present
                    if expense.has_cashback:
                        initial_data["has_cashback"] = True
                        initial_data["cashback_type"] = expense.cashback_type
                        initial_data["cashback_value"] = expense.cashback_value

                    # Handle shared expense data
                    if hasattr(expense, "shared_details") and expense.shared_details:
                        shared = expense.shared_details
                        participants_data = []

                        for participant in shared.participants.all():
                            # Find the share for this participant
                            share = shared.shares.filter(
                                participant=participant
                            ).first()
                            participants_data.append(
                                {
                                    "id": participant.id,
                                    "name": participant.name,
                                    "is_user": participant.is_user,
                                    "is_payer": participant.is_payer,
                                    "amount": str(share.amount) if share else "0",
                                }
                            )

                        # Find payer
                        payer = shared.participants.filter(is_payer=True).first()

                        copy_shared_data = {
                            "participants_json": json.dumps(participants_data),
                            "payer_id": payer.name if payer else "You",
                        }

                        # Set initial values for hidden fields
                        initial_data["participants_json"] = copy_shared_data[
                            "participants_json"
                        ]
                        initial_data["payer_id"] = copy_shared_data["payer_id"]
                        initial_data["expense_type"] = "shared"

                except Expense.DoesNotExist:
                    messages.error(request, "Expense not found.")

            # Use single form for regular/shared expense entry
            form = ExpenseForm(
                user=request.user, initial=initial_data if copy_expense_id else None
            )
            next_url = request.GET.get("next", "")

            # Pass additional context
            context = {
                "form": form,
                "next_url": next_url,
                "is_copy": bool(copy_expense_id),
            }

            if copy_shared_data:
                context["copy_shared_data"] = copy_shared_data

            return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        from django.db import transaction
        from .models import SharedExpense, SharedExpenseParticipant, Share, Friend
        from decimal import Decimal

        # Check if this is a bulk submission (formset) or single form
        bulk_mode = "form-TOTAL_FORMS" in request.POST

        if bulk_mode:
            # Handle formset submission (bulk mode)
            ExpenseFormSet = modelformset_factory(
                Expense, form=ExpenseForm, extra=1, can_delete=True
            )
            formset = ExpenseFormSet(request.POST, form_kwargs={"user": request.user})

            if formset.is_valid():
                try:
                    instances = formset.save(commit=False)

                    for instance in instances:
                        instance.user = request.user

                        # Get the form for this instance to access expense_type
                        form_index = instances.index(instance)
                        form = formset.forms[form_index]

                        expense_type = form.data.get(
                            f"{formset.prefix}-{form_index}-expense_type", "personal"
                        )

                        # Save the base expense first
                        instance.save()

                        # If this is a shared expense, create related records
                        if expense_type == "shared":
                            participants_json = form.cleaned_data.get(
                                "participants_json"
                            )
                            payer_id = form.cleaned_data.get("payer_id")

                            if participants_json and payer_id:
                                with transaction.atomic():
                                    # Parse participants data
                                    participants_data = json.loads(participants_json)

                                    # Find payer data
                                    payer_name = payer_id.strip()

                                    # Check if payer is "You" (the logged-in user)
                                    payer_is_user = payer_name == "You"

                                    # Check if user is in participants
                                    user_in_participants = any(
                                        p.get("is_user", False)
                                        for p in participants_data
                                    )

                                    # Create SharedExpense first (no payer FK anymore)
                                    shared_expense = SharedExpense.objects.create(
                                        expense=instance
                                    )

                                    # Create all SharedExpenseParticipant records
                                    participant_map = {}

                                    for participant_data in participants_data:
                                        participant_name = participant_data.get(
                                            "name", ""
                                        ).strip()
                                        is_user = participant_data.get("is_user", False)
                                        is_payer = (is_user and payer_is_user) or (
                                            participant_name == payer_name
                                        )

                                        # Get or create Friend from master table (not for user)
                                        friend = None
                                        if not is_user:
                                            friend, _ = Friend.objects.get_or_create(
                                                user=request.user, name=participant_name
                                            )

                                        participant = (
                                            SharedExpenseParticipant.objects.create(
                                                shared_expense=shared_expense,
                                                friend=friend,
                                                is_user=is_user,
                                                is_payer=is_payer,
                                            )
                                        )

                                        participant_map[participant_name] = participant

                                    # If payer is user but not in participants, create a participant record for them as payer only
                                    if payer_is_user and not user_in_participants:
                                        payer_participant = (
                                            SharedExpenseParticipant.objects.create(
                                                shared_expense=shared_expense,
                                                friend=None,
                                                is_user=True,
                                                is_payer=True,
                                            )
                                        )
                                        participant_map["You"] = payer_participant

                                    # Create Share records (only for participants with share amounts)
                                    for participant_data in participants_data:
                                        participant_name = participant_data.get(
                                            "name", ""
                                        ).strip()
                                        share_amount = participant_data.get(
                                            "share_amount"
                                        )

                                        if (
                                            share_amount is not None
                                            and share_amount != ""
                                        ):
                                            participant = participant_map[
                                                participant_name
                                            ]

                                            Share.objects.create(
                                                shared_expense=shared_expense,
                                                participant=participant,
                                                amount=Decimal(str(share_amount)),
                                            )

                    next_url = request.POST.get("next") or request.GET.get("next")
                    if next_url:
                        return redirect(next_url)
                    return redirect("expense-list")

                except IntegrityError as e:
                    messages.error(
                        request, f"This expense entry already exists: {str(e)}"
                    )
                    return render(request, self.template_name, {"formset": formset})
                except (ValueError, KeyError, json.JSONDecodeError) as e:
                    messages.error(request, f"Error creating shared expense: {str(e)}")
                    return render(request, self.template_name, {"formset": formset})

            return render(request, self.template_name, {"formset": formset})

        else:
            # Handle single form submission
            form = ExpenseForm(request.POST, user=request.user)

            if form.is_valid():
                try:
                    with transaction.atomic():
                        # Create the base expense
                        expense = form.save(commit=False)
                        expense.user = request.user
                        has_cashback = request.POST.get("has_cashback", "off")

                        if has_cashback == "on":
                            cashback_type = request.POST.get("cashback_type", "FIXED")
                            cashback_value = request.POST.get("cashback_value", "0")
                            expense.has_cashback = True
                            expense.cashback_type = cashback_type
                            expense.cashback_value = cashback_value
                            expense.amount = Decimal(expense.amount) - Decimal(expense.cashback_amount)

                        expense.save()

                        # Update account balances
                        expense.apply_payment_impact()

                        # Check if this is a shared expense
                        expense_type = request.POST.get("expense_type", "personal")

                        if expense_type == "shared":
                            participants_json = form.cleaned_data.get(
                                "participants_json"
                            )
                            payer_id = form.cleaned_data.get("payer_id")

                            if participants_json and payer_id:
                                # Parse participants data
                                participants_data = json.loads(participants_json)

                                # Find payer name
                                payer_name = payer_id.strip()

                                # Check if payer is "You" (the logged-in user)
                                payer_is_user = payer_name == "You"

                                # Check if user is in participants
                                user_in_participants = any(
                                    p.get("is_user", False) for p in participants_data
                                )

                                # Create SharedExpense first (no payer FK anymore)
                                shared_expense = SharedExpense.objects.create(
                                    expense=expense
                                )

                                # Create all SharedExpenseParticipant records
                                participant_map = {}

                                for participant_data in participants_data:
                                    participant_name = participant_data.get(
                                        "name", ""
                                    ).strip()
                                    is_user = participant_data.get("is_user", False)
                                    is_payer = (is_user and payer_is_user) or (
                                        participant_name == payer_name
                                    )

                                    # Get or create Friend from master table (not for user)
                                    friend = None
                                    if not is_user:
                                        friend, _ = Friend.objects.get_or_create(
                                            user=request.user, name=participant_name
                                        )

                                    participant = (
                                        SharedExpenseParticipant.objects.create(
                                            shared_expense=shared_expense,
                                            friend=friend,
                                            is_user=is_user,
                                            is_payer=is_payer,
                                        )
                                    )

                                    participant_map[participant_name] = participant

                                # If payer is user but not in participants, create a participant record for them as payer only
                                if payer_is_user and not user_in_participants:
                                    payer_participant = (
                                        SharedExpenseParticipant.objects.create(
                                            shared_expense=shared_expense,
                                            friend=None,
                                            is_user=True,
                                            is_payer=True,
                                        )
                                    )
                                    participant_map["You"] = payer_participant

                                # Create Share records (only for participants with share amounts)
                                for participant_data in participants_data:
                                    participant_name = participant_data.get(
                                        "name", ""
                                    ).strip()
                                    share_amount = participant_data.get("share_amount")

                                    if share_amount is not None and share_amount != "":
                                        participant = participant_map[participant_name]

                                        Share.objects.create(
                                            shared_expense=shared_expense,
                                            participant=participant,
                                            amount=Decimal(str(share_amount)),
                                        )

                        messages.success(request, "Expense created successfully!")

                        next_url = request.POST.get("next") or request.GET.get("next")
                        if next_url:
                            return redirect(next_url)
                        return redirect("expense-list")

                except IntegrityError as e:
                    messages.error(
                        request, f"This expense entry already exists: {str(e)}"
                    )
                    return render(request, self.template_name, {"form": form})
                except (ValueError, KeyError, json.JSONDecodeError) as e:
                    messages.error(request, f"Error creating shared expense: {str(e)}")
                    return render(request, self.template_name, {"form": form})

            return render(request, self.template_name, {"form": form})


class ExpenseUpdateView(LoginRequiredMixin, generic.UpdateView):
    model = Expense
    form_class = ExpenseForm
    template_name = 'expenses/expense_form.html'
    success_url = reverse_lazy('expense-list')

    def get_success_url(self):
        next_url = self.request.POST.get('next') or self.request.GET.get('next')
        if next_url:
            return next_url
        return super().get_success_url()

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['next_url'] = self.request.POST.get('next') or self.request.GET.get('next') or ''
        return context

    def get_initial(self):
        """Populate form with existing shared expense data if it exists."""
        initial = super().get_initial()
        expense = self.object

        try:
            # Check if this expense has shared details
            shared_expense = expense.shared_details

            # Set expense type to shared
            initial["expense_type"] = "shared"

            # Build participants list from existing data
            participants = []
            payer_name = None

            for participant in shared_expense.participants.all():
                # Get the participant's share
                share = shared_expense.shares.filter(participant=participant).first()
                share_amount = str(share.amount) if share else ""

                participant_data = {
                    "name": participant.name,
                    "is_user": participant.is_user,
                    "share_amount": share_amount,
                }
                participants.append(participant_data)

                # Track who paid
                if participant.is_payer:
                    payer_name = participant.name

            # Set the initial values for hidden fields
            initial["participants_json"] = json.dumps(participants)
            initial["payer_id"] = payer_name

        except Exception as e:
            # Not a shared expense or error loading data - use defaults
            # This is expected for personal expenses
            pass

        return initial

    def form_valid(self, form):
        try:
            # Get the old expense to restore balances
            old_expense = self.get_object()
            # Restore old balances first
            old_expense.revert_payment_impact()

            # Save the updated expense
            response = super().form_valid(form)

            # Deduct new balances
            new_expense = self.object
            new_expense.apply_payment_impact()

            return response
        except IntegrityError:
            messages.error(self.request, "This expense entry already exists.")
            return self.form_invalid(form)

    def get_queryset(self):
        # Ensure user can only edit their own expenses
        return Expense.objects.filter(user=self.request.user)

class ExpenseBulkDeleteView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        expense_ids = request.POST.getlist('expense_ids')
        if not expense_ids:
            messages.error(request, 'No expenses selected for deletion.')
            return redirect('expense-list')
            
        # Filter by IDs and ensuring they belong to the current user for security
        expenses_to_delete = Expense.objects.filter(id__in=expense_ids, user=request.user)
        
        # Restore balances for each expense before deleting
        for expense in expenses_to_delete:
            expense.revert_payment_impact()
        
        deleted_count = expenses_to_delete.count()
        
        if deleted_count > 0:
            expenses_to_delete.delete()
            messages.success(request, f'{deleted_count} expenses deleted successfully.')
        else:
            messages.warning(request, 'No valid expenses found to delete.')
            
        return redirect('expense-list')

class ExpenseDeleteView(LoginRequiredMixin, generic.DeleteView):
    model = Expense
    template_name = 'expenses/expense_confirm_delete.html'
    success_url = reverse_lazy('expense-list')

    def get_queryset(self):
        return Expense.objects.filter(user=self.request.user)

    def form_valid(self, form):
        """
        Browser delete requests hit POST -> form_valid(), not delete().
        Restore account/card balance before deleting the expense.
        """
        with transaction.atomic():
            self.object = self.get_object()
            self.object.revert_payment_impact()
            return super().form_valid(form)

    def delete(self, request, *args, **kwargs):
        """Override delete to restore account balances."""
        expense = self.get_object()
        
        # Restore balances before deleting
        expense.revert_payment_impact()
        
        return super().delete(request, *args, **kwargs)

class CategoryListView(LoginRequiredMixin, generic.ListView):
    model = Category
    template_name = 'expenses/category_list.html'
    context_object_name = 'categories'
    paginate_by = 10

    def get_queryset(self):
        queryset = Category.objects.filter(user=self.request.user).order_by('name')
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(name__icontains=search_query)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        return context

class CategoryCreateView(LoginRequiredMixin, generic.CreateView):
    model = Category
    fields = ['name', 'limit']
    template_name = 'expenses/category_form.html'
    success_url = reverse_lazy('category-list')

    def form_valid(self, form):
        try:
            form.instance.user = self.request.user
            return super().form_valid(form)
        except IntegrityError:
            messages.error(self.request, "This category already exists.")
            return self.form_invalid(form)

class CategoryUpdateView(LoginRequiredMixin, generic.UpdateView):
    model = Category
    fields = ['name', 'limit']
    template_name = 'expenses/category_form.html'
    success_url = reverse_lazy('category-list')

    def get_success_url(self):
        next_url = self.request.POST.get('next') or self.request.GET.get('next')
        if next_url:
            return next_url
        return super().get_success_url()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['next_url'] = self.request.POST.get('next') or self.request.GET.get('next') or ''
        return context

    def get_queryset(self):
        return Category.objects.filter(user=self.request.user)
    
    def form_valid(self, form):
        try:
            # Store old name to update related expenses
            old_name = self.get_object().name
            response = super().form_valid(form)
            new_name = self.object.name
            
            if old_name != new_name:
                Expense.objects.filter(user=self.request.user, category=old_name).update(category=new_name)
                
            return response
        except IntegrityError:
            messages.error(self.request, "This category already exists.")
            return self.form_invalid(form)

class CategoryDeleteView(LoginRequiredMixin, generic.DeleteView):
    model = Category
    template_name = 'expenses/category_confirm_delete.html'
    success_url = reverse_lazy('category-list')

    def get_queryset(self):
        return Category.objects.filter(user=self.request.user)

@login_required
def export_expenses(request):
    """
    Export expenses to CSV based on current filters.
    """
    # Check Limits
    if not request.user.profile.is_plus:
        messages.error(request, "Export is available on Plus and Pro plans.")
        return redirect('pricing')

    expenses = Expense.objects.filter(user=request.user).order_by('-date')

    # Filter Logic
    selected_years = request.GET.getlist('year')
    selected_months = request.GET.getlist('month')
    selected_categories = request.GET.getlist('category')
    search_query = request.GET.get('search')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Remove empty strings
    selected_years = [y for y in selected_years if y]
    selected_months = [m for m in selected_months if m]
    selected_categories = [c for c in selected_categories if c]

    # Date Range Logic (Precedence over Year/Month)
    if start_date or end_date:
        if start_date:
            expenses = expenses.filter(date__gte=start_date)
        if end_date:
            expenses = expenses.filter(date__lte=end_date)
    else:
        # Check if any specific filter is active
        has_active_filters = (
            selected_years or 
            selected_months or 
            search_query
        )
        
        # If no year/month/search filters, default to current month/year
        if not has_active_filters:
            selected_years = [str(datetime.now().year)]
            selected_months = [str(datetime.now().month)]
        
        if selected_years:
            expenses = expenses.filter(date__year__in=selected_years)
        
        if selected_months:
            expenses = expenses.filter(date__month__in=selected_months)

    if selected_categories:
        expenses = expenses.filter(category__in=selected_categories)
    if search_query:
        expenses = expenses.filter(description__icontains=search_query)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="expenses.csv"'

    writer = csv.writer(response)
    writer.writerow(['Date', 'Category', 'Description', 'Amount'])

    for expense in expenses:
        writer.writerow([expense.date, expense.category, expense.description, expense.amount])

    return response

# --------------------
# Income Views
# --------------------

class IncomeListView(LoginRequiredMixin, RecurringTransactionMixin, ListView):
    model = Income
    template_name = 'expenses/income_list.html'
    context_object_name = 'incomes'
    paginate_by = 20

    def get_queryset(self):
        queryset = Income.objects.filter(user=self.request.user).order_by('-date')
        
        # Default dates (Current Year)
        today = timezone.localdate()
        default_start = today.replace(month=1, day=1)
        default_end = today.replace(month=12, day=31)

        # Date Filter
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        source = self.request.GET.get('source')

        # Check if we have ANY filter params. If not, apply default dates.
        if not date_from and not date_to and not source:
             self.date_from = default_start.isoformat()
             self.date_to = default_end.isoformat()
             queryset = queryset.filter(date__gte=default_start, date__lte=default_end)
        else:
            # We have some filters (or user explicitly cleared them? - tricky part about "reset")
            # If user wants to "clear" filters, they usually submit empty strings.
            # But the requirement says "default start date...". Usually implies initial load.
            if date_from:
                queryset = queryset.filter(date__gte=date_from)
                self.date_from = date_from
            else:
                self.date_from = ''
            
            if date_to:
                queryset = queryset.filter(date__lte=date_to)
                self.date_to = date_to
            else:
                self.date_to = ''

        # Source Filter
        if source:
            queryset = queryset.filter(source__icontains=source)
            
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Calculate stats for the filtered queryset
        filtered_queryset = self.object_list
        context['filtered_count'] = filtered_queryset.count()
        context['filtered_amount'] = filtered_queryset.aggregate(Sum('amount'))['amount__sum'] or 0
        
        context['filter_form'] = {
            'date_from': getattr(self, 'date_from', ''),
            'date_to': getattr(self, 'date_to', ''),
            'source': self.request.GET.get('source', ''),
        }
        return context

class IncomeCreateView(LoginRequiredMixin, generic.CreateView):
    model = Income
    form_class = IncomeForm
    template_name = 'expenses/income_form.html'
    success_url = reverse_lazy('income-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        try:
            form.instance.user = self.request.user
            return super().form_valid(form)
        except IntegrityError:
            messages.error(self.request, "This income entry already exists.")
            return self.form_invalid(form)

    def get_success_url(self):
        next_url = self.request.POST.get('next') or self.request.GET.get('next')
        if next_url:
            return next_url
        return super().get_success_url()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['next_url'] = self.request.POST.get('next') or self.request.GET.get('next') or ''
        return context


class IncomeUpdateView(LoginRequiredMixin, generic.UpdateView):
    model = Income
    form_class = IncomeForm
    template_name = 'expenses/income_form.html'
    success_url = reverse_lazy('income-list')

    def get_success_url(self):
        next_url = self.request.POST.get('next') or self.request.GET.get('next')
        if next_url:
            return next_url
        return super().get_success_url()

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['next_url'] = self.request.POST.get('next') or self.request.GET.get('next') or ''
        return context

    def get_queryset(self):
        return Income.objects.filter(user=self.request.user)

    def form_valid(self, form):
        try:
            return super().form_valid(form)
        except IntegrityError:
            messages.error(self.request, "This income entry already exists.")
            return self.form_invalid(form)

class IncomeDeleteView(LoginRequiredMixin, generic.DeleteView):
    model = Income
    template_name = 'expenses/income_confirm_delete.html'
    success_url = reverse_lazy('income-list')

    def get_queryset(self):
        return Income.objects.filter(user=self.request.user)



class CalendarView(LoginRequiredMixin, RecurringTransactionMixin, TemplateView):
    template_name = 'expenses/calendar.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = datetime.now()
        
        # Get year/month from URL or default to current
        year = self.kwargs.get('year', today.year)
        month = self.kwargs.get('month', today.month)
        
        # Validate year/month
        try:
            year = int(year)
            month = int(month)
            if month < 1 or month > 12:
                raise ValueError
        except ValueError:
            year = today.year
            month = today.month

        # Calculate prev/next month for navigation
        if month == 1:
            prev_month_date = date(year - 1, 12, 1)
        else:
            prev_month_date = date(year, month - 1, 1)
            
        if month == 12:
            next_month_date = date(year + 1, 1, 1)
        else:
            next_month_date = date(year, month + 1, 1)

        # Get search query
        search_query = self.request.GET.get('search', '')

        # Base filters
        expense_filters = Q(user=self.request.user, date__year=year, date__month=month)
        income_filters = Q(user=self.request.user, date__year=year, date__month=month)
        
        if search_query:
            # Filter expenses by description or category
            expense_filters &= (Q(description__icontains=search_query) | Q(category__icontains=search_query))
            # Filter income by source or description
            income_filters &= (Q(source__icontains=search_query) | Q(description__icontains=search_query))

        # Get Expense and Income Data for the month
        expenses = Expense.objects.filter(expense_filters).values('date').annotate(total=Sum('amount'))
        
        incomes = Income.objects.filter(income_filters).values('date').annotate(total=Sum('amount'))
        
        # Map data for easy lookup by day
        # Keys are integers (day of month)
        expense_map = {e['date'].day: e['total'] for e in expenses}
        income_map = {i['date'].day: i['total'] for i in incomes}
        
        # Build Calendar Grid
        cal = calendar.Calendar(firstweekday=6) # Start on Sunday
        month_days = cal.monthdayscalendar(year, month)
        
        # Transform into a list of weeks, where each day is an object
        calendar_data = []
        for week in month_days:
            week_data = []
            for day in week:
                if day == 0:
                    week_data.append(None) # Empty slot
                else:
                    week_data.append({
                        'day': day,
                        'income': income_map.get(day, 0),
                        'expense': expense_map.get(day, 0),
                    })
            calendar_data.append(week_data)
        
        
        # Calculate totals for the month to show net savings
        total_monthly_expense = sum(item['total'] for item in expenses) or 0
        total_monthly_income = sum(item['total'] for item in incomes) or 0
        month_net_savings = total_monthly_income - total_monthly_expense

        context['calendar_data'] = calendar_data
        context['current_year'] = year
        context['current_month'] = month
        context['month_name'] = calendar.month_name[month]
        context['month_net_savings'] = month_net_savings
        context['prev_year'] = prev_month_date.year
        context['prev_month'] = prev_month_date.month
        context['next_year'] = next_month_date.year
        context['next_month'] = next_month_date.month
        context['search_query'] = search_query
        
        return context


class BudgetDashboardView(LoginRequiredMixin, RecurringTransactionMixin, TemplateView):
    template_name = 'expenses/budget_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        today = date.today()
        
        month_param = self.request.GET.get('month')
        year_param = self.request.GET.get('year')
        
        month = int(month_param) if month_param else today.month
        year = int(year_param) if year_param else today.year
        
        # Ensure context variables for filters are correct
        context['current_month'] = month
        context['current_year'] = year
        
        categories = Category.objects.filter(user=user)
        budget_data = []
        
        total_budget = 0
        categorized_spent = 0
        
        # Calculate total spending across ALL expenses for the month
        grand_total_spent = Expense.objects.filter(
            user=user,
            date__year=year,
            date__month=month
        ).aggregate(Total=Sum('amount'))['Total'] or 0

        for category in categories:
            spent = Expense.objects.filter(
                user=user,
                category=category.name,
                date__year=year,
                date__month=month
            ).aggregate(Total=Sum('amount'))['Total'] or 0
            
            percentage = (spent / category.limit * 100) if category.limit and category.limit > 0 else 0
            
            budget_data.append({
                'category': category,
                'spent': spent,
                'limit': category.limit,
                'percentage': min(percentage, 100),
                'actual_percentage': percentage,
                'remaining': (category.limit - spent) if category.limit and spent <= category.limit else 0,
                'over_budget': (spent - category.limit) if category.limit and spent > category.limit else 0
            })
            
            if category.limit:
                total_budget += category.limit
            categorized_spent += spent
            
        context.update({
            'budget_data': budget_data,
            'total_budget': total_budget,
            'total_spent': grand_total_spent,
            'total_remaining': (total_budget - grand_total_spent) if total_budget > grand_total_spent else 0,
            'over_budget_amount': (grand_total_spent - total_budget) if grand_total_spent > total_budget else 0,
            'total_percentage': min((grand_total_spent / total_budget * 100), 100) if total_budget else 0,
            'actual_total_percentage': (grand_total_spent / total_budget * 100) if total_budget else 0,
            'month_name': date(year, month, 1).strftime('%B'),
        })

        # MoM Calculation for Budget Dashboard
        if month == 1:
            prev_month = 12
            prev_year = year - 1
        else:
            prev_month = month - 1
            prev_year = year

        prev_spent = Expense.objects.filter(
            user=user,
            date__year=prev_year,
            date__month=prev_month
        ).aggregate(Total=Sum('amount'))['Total'] or 0

        if prev_spent > 0:
            context['spent_mom_pct'] = ((grand_total_spent - prev_spent) / prev_spent) * 100
            context['spent_mom_pct_abs'] = abs(context['spent_mom_pct'])
        else:
            context['spent_mom_pct'] = None
            context['spent_mom_pct_abs'] = None

        context.update({
            'current_month': month,
            'current_year': year,
            'months': [(i, calendar.month_name[i]) for i in range(1, 13)],
            'years': range(today.year - 2, today.year + 2),
        })
        return context

# --------------------
# Cash Credit Views
# Cash received as credit from someone to bank account or given to someone as credit (on repayment enter the amount repaid and )
# --------------------





# --------------------
# Recurring Transaction Views
# --------------------

class RecurringTransactionListView(LoginRequiredMixin, ListView):
    model = RecurringTransaction
    template_name = 'expenses/recurring_transaction_list.html'
    context_object_name = 'recurring_transactions'
    filter_expenses_only = True

    def get_queryset(self):
        queryset = RecurringTransaction.objects.filter(user=self.request.user)
        if self.filter_expenses_only:
            queryset = queryset.filter(transaction_type='EXPENSE')
        queryset = queryset.order_by('-created_at')
        
        # Filter by Category
        categories = self.request.GET.getlist('category')
        if categories:
            queryset = queryset.filter(category__in=categories)
            
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        all_transactions = self.object_list
        today = date.today()
        
        # Categories for filter
        user_transactions = RecurringTransaction.objects.filter(user=self.request.user)
        categories = user_transactions.values_list('category', flat=True).distinct().order_by('category')
        # Filter out None/Empty if any
        categories = [c for c in categories if c]
        
        context['categories'] = categories
        context['selected_categories'] = self.request.GET.getlist('category')
        
        # Split into Active and Cancelled
        active_subs = [t for t in all_transactions if t.is_active]
        cancelled_subs = [t for t in all_transactions if not t.is_active]
        
        # Calculate Totals (Monthly & Yearly)
        total_monthly = 0
        total_yearly = 0
        
        for sub in active_subs:
            amount = sub.amount
            if sub.frequency == 'DAILY':
                total_monthly += amount * 30
                total_yearly += amount * 365
            elif sub.frequency == 'WEEKLY':
                total_monthly += amount * 4
                total_yearly += amount * 52
            elif sub.frequency == 'MONTHLY':
                total_monthly += amount
                total_yearly += amount * 12
            elif sub.frequency == 'YEARLY':
                total_monthly += amount / 12
                total_yearly += amount

        # Identify "Renewing Soon" (This Month)
        renewing_soon = []
        renewals_count = 0
        
        # Helper to find next date relative to today
        for sub in active_subs:
            # Calculate next occurrence
            next_date = sub.start_date
            
            # For simpler logic, we reset the year/month to current to check basic interval
            # But for accurate "days until", we need better logic:
            
            if sub.frequency == 'DAILY':
                next_date = today + timedelta(days=1)
            elif sub.frequency == 'WEEKLY':
                # Find days ahead
                days_ahead = (sub.start_date.weekday() - today.weekday()) % 7
                if days_ahead == 0 and today > sub.start_date: # if today is the day, but older start
                     days_ahead = 7
                elif days_ahead == 0 and today == sub.start_date: # exact match today
                     days_ahead = 0
                else: 
                     # If start_date was future, we wait. If past, we find next.
                     # Simplified: just next occurrence of that weekday
                     if days_ahead <= 0: days_ahead += 7
                
                # Correction: Standard logic to find next matching weekday
                days_ahead = (sub.start_date.weekday() - today.weekday()) 
                if days_ahead <= 0: # Target day already happened this week or is today
                    days_ahead += 7
                next_date = today + timedelta(days=days_ahead)
                
            elif sub.frequency == 'MONTHLY':
                # Occurs on sub.start_date.day every month
                # If today.day > start_date.day, it's next month.
                # If today.day <= start_date.day, it's this month.
                try:
                    if today.day > sub.start_date.day:
                        # Next month
                        month = today.month + 1
                        year = today.year
                        if month > 12:
                            month = 1
                            year += 1
                        next_date = date(year, month, sub.start_date.day)
                    else:
                        # This month
                        next_date = date(today.year, today.month, sub.start_date.day)
                except ValueError: 
                    # Handle end of month issues (e.g. 31st) - simplified to 1st of next-next month
                    next_date = (today.replace(day=1) + timedelta(days=32)).replace(day=1)

            elif sub.frequency == 'YEARLY':
                try:
                    this_year_date = date(today.year, sub.start_date.month, sub.start_date.day)
                    if today > this_year_date:
                        next_date = date(today.year + 1, sub.start_date.month, sub.start_date.day)
                    else:
                        next_date = this_year_date
                except ValueError:
                    next_date = date(today.year, 2, 28)

            # Annotate object
            sub.annotated_next_date = next_date
            sub.annotated_days_until = (next_date - today).days
            
            # Determine urgency
            is_renewing = False
            if sub.transaction_type == 'EXPENSE':
                if sub.annotated_days_until <= 30: # Show mostly anything coming up soon
                     is_renewing = True
            
            if is_renewing:
                renewing_soon.append(sub)
                renewals_count += 1
            
            # Sort renewing soon by days until
            renewing_soon.sort(key=lambda x: x.annotated_days_until)

        context.update({
            'active_subs': active_subs,
            'cancelled_subs': cancelled_subs,
            'renewing_soon': renewing_soon,
            'renewals_count': renewals_count,
            'total_monthly_cost': total_monthly,
            'total_yearly_cost': total_yearly,
        })
        return context

class RecurringTransactionManageView(RecurringTransactionListView):
    template_name = 'expenses/recurring_transaction_manage.html'
    filter_expenses_only = False

class RecurringTransactionCreateView(LoginRequiredMixin, CreateView):
    model = RecurringTransaction
    form_class = RecurringTransactionForm
    template_name = 'expenses/recurring_transaction_form.html'
    success_url = reverse_lazy('recurring-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        # Check Limits
        current_count = RecurringTransaction.objects.filter(user=self.request.user, is_active=True).count()
        limit = 0 # Free
        if self.request.user.profile.is_plus:
            limit = 3
        if self.request.user.profile.is_pro:
            limit = float('inf')

        if current_count >= limit:
             messages.error(self.request, f"Recurring Transaction limit reached ({limit}). Please upgrade.")
             return redirect('pricing')
             
        form.instance.user = self.request.user
        messages.success(self.request, 'Recurring transaction created successfully.')
        return super().form_valid(form)

    def get_success_url(self):
        next_url = self.request.POST.get('next') or self.request.GET.get('next')
        if next_url:
            return next_url
        return super().get_success_url()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['next_url'] = self.request.POST.get('next') or self.request.GET.get('next') or ''
        return context

    def get_success_url(self):
        next_url = self.request.POST.get('next') or self.request.GET.get('next')
        if next_url:
            return next_url
        return super().get_success_url()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['next_url'] = self.request.POST.get('next') or self.request.GET.get('next') or ''
        return context

class RecurringTransactionUpdateView(LoginRequiredMixin, UpdateView):
    model = RecurringTransaction
    form_class = RecurringTransactionForm
    template_name = 'expenses/recurring_transaction_form.html'
    success_url = reverse_lazy('recurring-list')

    def get_success_url(self):
        next_url = self.request.POST.get('next') or self.request.GET.get('next')
        if next_url:
            return next_url
        return super().get_success_url()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['next_url'] = self.request.POST.get('next') or self.request.GET.get('next') or ''
        return context

    def get_queryset(self):
        return RecurringTransaction.objects.filter(user=self.request.user)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        # Check if is_active changed from True to False
        if 'is_active' in form.changed_data and not form.cleaned_data['is_active']:
            # Cancellation detected
            obj = self.get_object() # get current object
            # Calculate yearly saving
            amount = obj.amount
            if obj.frequency == 'DAILY':
                yearly_saving = amount * 365
            elif obj.frequency == 'WEEKLY':
                yearly_saving = amount * 52
            elif obj.frequency == 'MONTHLY':
                yearly_saving = amount * 12
            else: # YEARLY
                yearly_saving = amount
            
            # Assuming currency symbol is available in request or we use generic. 
            # We can use the profile currency if available, or just a generic prompt. 
            # User request used '₹', but code uses {{ currency_symbol }} in template.
            # We'll try to fetch user currency or default.
            currency = '₹'
            if hasattr(self.request.user, 'userprofile'):
                currency = self.request.user.userprofile.currency
                
            messages.success(self.request, f"You just saved {currency}{yearly_saving:,.0f}/year 🎉")
        else:
            messages.success(self.request, 'Recurring transaction updated successfully.')
            
        return super().form_valid(form)

class RecurringTransactionDeleteView(LoginRequiredMixin, DeleteView):
    model = RecurringTransaction
    template_name = 'expenses/recurring_transaction_confirm_delete.html' # Added template_name for consistency
    success_url = reverse_lazy('recurring-list')

    def get_queryset(self):
        return RecurringTransaction.objects.filter(user=self.request.user)

    def form_valid(self, form):
        # Calculate savings
        obj = self.object
        amount = obj.amount
        if obj.frequency == 'DAILY':
            yearly_saving = amount * 365
        elif obj.frequency == 'WEEKLY':
            yearly_saving = amount * 52
        elif obj.frequency == 'MONTHLY':
            yearly_saving = amount * 12
        else: # YEARLY
            yearly_saving = amount
            
        currency = '₹'
        if hasattr(self.request.user, 'userprofile'):
            currency = self.request.user.userprofile.currency
            
        messages.success(self.request, f"You just saved {currency}{yearly_saving:,.0f}/year 🎉")
        return super().form_valid(form)


# --------------------
# Cash Credit Views
# --------------------


class CashCreditListView(LoginRequiredMixin, ListView):
    model = CashCredit
    template_name = "expenses/cash_credit_list.html"
    context_object_name = "cash_credits"
    paginate_by = 20

    def get_queryset(self):
        return CashCredit.objects.filter(user=self.request.user).select_related(
            "friend", "received_into_account", "given_from_account"
        ).order_by("-date", "-created_at")


class CashCreditCreateView(LoginRequiredMixin, CreateView):
    model = CashCredit
    form_class = CashCreditForm
    template_name = "expenses/cash_credit_form.html"
    success_url = reverse_lazy("cash-credit-list")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user
        return kwargs

    def form_valid(self, form):
        form.instance.user = self.request.user
        acc_given = form.cleaned_data.get("given_from_account")
        credit_type = form.cleaned_data.get("credit_type")
        total_amount = form.cleaned_data.get("total_amount")
        if credit_type == "lent" and acc_given and acc_given.balance < total_amount:
            messages.error(
                self.request,
                f"Not enough balance in {acc_given.name}. Current: {acc_given.balance}.",
            )
            return self.render_to_response(self.get_context_data(form=form))
        obj = form.save()
        self.object = obj  # required for get_success_url() which uses self.object
        acc_received = obj.received_into_account
        if obj.credit_type == "borrowed" and acc_received:
            acc_received.add(obj.total_amount)
        elif obj.credit_type == "lent" and acc_given:
            acc_given.deduct(obj.total_amount)
        messages.success(self.request, "Cash credit recorded successfully.")
        return redirect(self.get_success_url())


class CashCreditDetailView(LoginRequiredMixin, DetailView):
    model = CashCredit
    template_name = "expenses/cash_credit_detail.html"
    context_object_name = "credit"

    def get_queryset(self):
        return CashCredit.objects.filter(user=self.request.user).prefetch_related("repayments")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        credit = context.get("credit")
        context["repayment_form"] = CashCreditRepaymentForm(
            user=self.request.user, credit=credit
        )
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = CashCreditRepaymentForm(
            request.POST, user=request.user, credit=self.object
        )
        if form.is_valid():
            amount = form.cleaned_data["amount"]
            remaining_before = self.object.remaining
            if amount > remaining_before:
                messages.error(
                    request,
                    f"Repayment amount cannot exceed remaining amount ({remaining_before}).",
                )
                return redirect("cash-credit-detail", pk=self.object.pk)
            repayment = form.save(commit=False)
            repayment.cash_credit = self.object
            repayment.save()
            # LENT: friend paid me back -> add to received_into_account (mandatory)
            if self.object.credit_type == "lent" and repayment.received_into_account:
                repayment.received_into_account.add(repayment.amount)
            # BORROWED: I paid friend back -> deduct from paid_from_account (mandatory)
            elif self.object.credit_type == "borrowed" and repayment.paid_from_account:
                repayment.paid_from_account.deduct(repayment.amount)
            messages.success(request, f"Repayment of {amount} recorded.")
            return redirect("cash-credit-detail", pk=self.object.pk)
        context = self.get_context_data()
        context["repayment_form"] = form
        return self.render_to_response(context)


class CashCreditUpdateView(LoginRequiredMixin, UpdateView):
    model = CashCredit
    form_class = CashCreditForm
    template_name = "expenses/cash_credit_form.html"
    success_url = reverse_lazy("cash-credit-list")
    context_object_name = "credit"

    def get_queryset(self):
        return CashCredit.objects.filter(user=self.request.user)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user
        return kwargs

    def form_valid(self, form):
        obj = self.get_object()
        old_total = obj.total_amount
        old_received = obj.received_into_account
        old_given = obj.given_from_account
        new_total = form.cleaned_data["total_amount"]
        new_received = form.cleaned_data.get("received_into_account")
        new_given = form.cleaned_data.get("given_from_account")
        amount_repaid = obj.amount_repaid
        if new_total < amount_repaid:
            form.add_error(
                "total_amount",
                f"Total cannot be less than amount already repaid ({amount_repaid}).",
            )
            return self.form_invalid(form)
        # If account(s) changed, reverse full old and apply full new
        if obj.credit_type == "borrowed":
            if old_received != new_received:
                if old_received:
                    old_received.deduct(old_total)
                if new_received:
                    new_received.add(new_total)
            elif old_received and old_total != new_total:
                diff = new_total - old_total
                if diff > 0:
                    old_received.add(diff)
                else:
                    old_received.deduct(-diff)
        elif obj.credit_type == "lent":
            if old_given != new_given:
                if old_given:
                    old_given.add(old_total)
                if new_given:
                    if new_given.balance < new_total:
                        form.add_error(
                            "given_from_account",
                            f"Not enough balance in {new_given.name}.",
                        )
                        return self.form_invalid(form)
                    new_given.deduct(new_total)
            elif old_given and old_total != new_total:
                diff = new_total - old_total
                if diff > 0:
                    if old_given.balance < diff:
                        form.add_error(
                            "total_amount",
                            f"Not enough balance in {old_given.name} to increase by {diff}.",
                        )
                        return self.form_invalid(form)
                    old_given.deduct(diff)
                else:
                    old_given.add(-diff)
        messages.success(self.request, "Cash credit updated.")
        return super().form_valid(form)


class CashCreditDeleteView(LoginRequiredMixin, DeleteView):
    model = CashCredit
    template_name = "expenses/cash_credit_confirm_delete.html"
    success_url = reverse_lazy("cash-credit-list")
    context_object_name = "credit"

    def get_queryset(self):
        return CashCredit.objects.filter(user=self.request.user)

    def delete(self, request, *args, **kwargs):
        obj = self.get_object()
        # Reverse balance: BORROWED had added to account -> deduct; LENT had deducted -> add back
        acc_received = obj.received_into_account
        acc_given = obj.given_from_account
        repaid = obj.amount_repaid
        if obj.credit_type == "borrowed" and acc_received:
            acc_received.deduct(obj.total_amount)
        elif obj.credit_type == "lent" and acc_given:
            acc_given.add(obj.total_amount)
        # Repayments: reverse balance updates
        for r in obj.repayments.all():
            if r.received_into_account:
                r.received_into_account.deduct(r.amount)
            if r.paid_from_account:
                r.paid_from_account.add(r.amount)
        return super().delete(request, *args, **kwargs)


class CashCreditBulkDeleteView(LoginRequiredMixin, View):
    def post(self, request):
        ids = request.POST.getlist("ids")
        if not ids:
            messages.warning(request, "No items selected.")
            return redirect("cash-credit-list")
        qs = CashCredit.objects.filter(user=request.user, pk__in=ids)
        for obj in qs:
            acc_received = obj.received_into_account
            acc_given = obj.given_from_account
            if obj.credit_type == "borrowed" and acc_received:
                acc_received.deduct(obj.total_amount)
            elif obj.credit_type == "lent" and acc_given:
                acc_given.add(obj.total_amount)
            for r in obj.repayments.all():
                if r.received_into_account:
                    r.received_into_account.deduct(r.amount)
                if r.paid_from_account:
                    r.paid_from_account.add(r.amount)
        count = qs.count()
        qs.delete()
        messages.success(request, f"{count} cash credit(s) deleted.")
        return redirect("cash-credit-list")


class AccountDeleteView(LoginRequiredMixin, DeleteView):
    model = User
    success_url = reverse_lazy('landing')
    template_name = 'expenses/account_confirm_delete.html'

    def get_object(self, queryset=None):
        return self.request.user

    def delete(self, request, *args, **kwargs):
        user = self.get_object()
        logout(request) # Log out before deleting
        user.delete()
        messages.success(request, "Your account has been deleted successfully.")
        return redirect(self.success_url)

class CurrencyUpdateView(LoginRequiredMixin, UpdateView):
    model = UserProfile
    fields = ['currency']
    template_name = 'expenses/currency_settings.html'
    success_url = reverse_lazy('currency-settings')

    def get_object(self, queryset=None):
        profile, created = UserProfile.objects.get_or_create(user=self.request.user)
        return profile

    def form_valid(self, form):
        messages.success(self.request, 'Currency preference updated successfully.')
        return super().form_valid(form)

class ProfileUpdateView(LoginRequiredMixin, UpdateView):
    model = User
    form_class = ProfileUpdateForm
    template_name = 'expenses/profile_settings.html'
    success_url = reverse_lazy('profile-settings')

    def get_object(self):
        return self.request.user

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Profile Settings'
        context['is_social_user'] = SocialAccount.objects.filter(user=self.request.user).exists()
        return context

    def form_valid(self, form):
        messages.success(self.request, "Profile updated successfully.")
        return super().form_valid(form)

def demo_login(request):
    """
    Logs in the read-only 'demo' user without password authentication.
    """
    # Clear any existing messages (e.g. from previous logout)
    list(messages.get_messages(request))

    try:
        user = User.objects.get(username='demo')
        # Manually set the backend to allow login without authentication
        login(request, user, backend='django.contrib.auth.backends.ModelBackend')
        messages.success(request, "🚀 Welcome to Demo Mode! Feel free to explore the app.")
        return redirect('home')
    except User.DoesNotExist:
        messages.error(request, "Demo user not setup. Please contact admin.")
        return redirect('account_login')

class PricingView(TemplateView):
    template_name = 'expenses/pricing.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['RAZORPAY_KEY_ID'] = settings.RAZORPAY_KEY_ID
        plans = SubscriptionPlan.objects.filter(is_active=True)
        context['plans'] = {p.tier: p for p in plans}
        return context

def ping(request):
    return HttpResponse("Pong", status=200)


class BalanceSummaryView(LoginRequiredMixin, TemplateView):
    """
    View to display balance summary showing who owes whom based on Friend master table.

    Requirements:
        - 7.1: Aggregate lent and borrowed amounts per friend per month
        - 7.2: Display total lent amount for each friend
        - 7.3: Display total borrowed amount for each friend
        - 7.4: Display net balance for each friend
        - 7.6: Allow filtering balance reports by specific months and years
        - 8.4: Calculate settlements across all time periods by default
        - 8.5: Allow filtering settlements by date range
    """

    template_name = "expenses/balance_summary.html"

    def get_context_data(self, **kwargs):
        from .services import BalanceCalculationService
        from .models import SharedExpense, Friend

        context = super().get_context_data(**kwargs)
        user = self.request.user

        # Get filter parameters
        month_param = self.request.GET.get("month")
        year_param = self.request.GET.get("year")
        friend_param = self.request.GET.get("friend")

        # Initialize date range variables
        start_date = None
        end_date = None
        filter_applied = False
        selected_friend_id = None

        # Process date filters if provided
        if month_param and year_param:
            try:
                month = int(month_param)
                year = int(year_param)

                # Validate month and year
                if 1 <= month <= 12 and 1900 <= year <= 9999:
                    # Calculate start and end dates for the month
                    start_date = date(year, month, 1)

                    # Get last day of the month
                    last_day = calendar.monthrange(year, month)[1]
                    end_date = date(year, month, last_day)

                    filter_applied = True
                    context["selected_month"] = month
                    context["selected_year"] = year
                    context["month_name"] = calendar.month_name[month]
            except (ValueError, TypeError):
                # Invalid month/year format - ignore and skip filter
                pass
        elif year_param:
            # Year only filter
            try:
                year = int(year_param)
                if 1900 <= year <= 9999:
                    start_date = date(year, 1, 1)
                    end_date = date(year, 12, 31)
                    filter_applied = True
                    context["selected_year"] = year
            except (ValueError, TypeError):
                # Invalid year format - ignore and skip filter
                pass

        # Process friend filter
        if friend_param:
            try:
                selected_friend_id = int(friend_param)
                filter_applied = True
                context["selected_friend_id"] = selected_friend_id
            except (ValueError, TypeError):
                # Invalid friend ID format - ignore and skip filter
                pass

        # Calculate balances using the service (now returns Friend-based data)
        balances = BalanceCalculationService.calculate_balances(
            user=user, start_date=start_date, end_date=end_date
        )

        # Get transaction details for each friend
        transactions_by_friend = BalanceCalculationService.get_transactions_by_friend(
            user=user, start_date=start_date, end_date=end_date
        )

        # Filter by friend if specified
        if selected_friend_id:
            balances = {k: v for k, v in balances.items() if k == selected_friend_id}
            transactions_by_friend = {
                k: v
                for k, v in transactions_by_friend.items()
                if k == selected_friend_id
            }

        # Prepare balance data for template
        # Separate into people who owe user (positive net) and people user owes (negative net)
        people_owe_user = []
        user_owes_people = []
        settled_people = []

        # user id
        for friend_id, balance_data in balances.items():
            if friend_id != user.id:
                net_balance = balance_data["net"]
                friend = balance_data.get("friend")

                balance_info = {
                    "id": friend_id,
                    "friend": friend,
                    "name": balance_data["name"],
                    "email": friend.email if friend else None,
                    "phone": friend.phone if friend else None,
                    "lent": balance_data["lent"],
                    "borrowed": balance_data["borrowed"],
                    "net": net_balance,
                    "net_abs": abs(net_balance),
                    "transactions": transactions_by_friend.get(friend_id, []),
                }

                if net_balance > 0:
                    # User lent more than borrowed - friend owes user
                    people_owe_user.append(balance_info)
                elif net_balance < 0:
                    # User borrowed more than lent - user owes friend
                    user_owes_people.append(balance_info)
                else:
                    # Net balance is zero - settled up
                    settled_people.append(balance_info)

        # Sort by absolute net balance (highest first)
        people_owe_user.sort(key=lambda x: x["net"], reverse=True)
        user_owes_people.sort(key=lambda x: x["net"])
        settled_people.sort(key=lambda x: x["name"])

        # Add to context
        context["people_owe_user"] = people_owe_user
        context["user_owes_people"] = user_owes_people
        context["settled_people"] = settled_people
        context["filter_applied"] = filter_applied

        # Calculate totals
        total_owed_to_user = sum(b["net"] for b in people_owe_user)
        total_user_owes = abs(sum(b["net"] for b in user_owes_people))
        overall_net = total_owed_to_user - total_user_owes

        context["total_owed_to_user"] = total_owed_to_user
        context["total_user_owes"] = total_user_owes
        context["overall_net"] = overall_net

        # Get all friends for the user (for potential friend management)
        # Show ALL friends, not just those with transactions
        context["all_friends"] = Friend.objects.filter(user=user).order_by("name")
        context["friends_count"] = context["all_friends"].count()

        # Provide month and year options for filter dropdowns
        today = date.today()
        context["months_list"] = [(i, calendar.month_name[i]) for i in range(1, 13)]

        # Get years from shared expenses
        shared_expenses = SharedExpense.objects.filter(
            expense__user=user
        ).select_related("expense")

        if shared_expenses.exists():
            years_set = set(se.expense.date.year for se in shared_expenses)
            years_set.add(today.year)
            context["years"] = sorted(years_set, reverse=True)
        else:
            context["years"] = [today.year]

        return context


class ContactView(View):
    template_name = 'contact.html'
    
    # Spam protection settings
    RATE_LIMIT_HOURLY = 3
    RATE_LIMIT_DAILY = 10
    MIN_MESSAGE_LENGTH = 10
    
    # Common spam patterns
    SPAM_KEYWORDS = [
        'precio', 'price check', 'buy now', 'click here', 'earn money',
        'viagra', 'casino', 'lottery', 'prize', 'congratulations',
        'limited offer', 'act now', 'online pharmacy', 'weight loss',
        'make money fast', 'work from home', 'investment opportunity',
        'hola, quería saber', 'please kindly', 'dear friend'
    ]
    
    # Disposable email domains
    DISPOSABLE_DOMAINS = [
        'tempmail.com', 'guerrillamail.com', '10minutemail.com',
        'throwaway.email', 'maildrop.cc', 'mailinator.com',
        'trashmail.com', 'yopmail.com', 'getnada.com'
    ]

    def get(self, request):
        form = ContactForm()
        return render(request, self.template_name, {'form': form})
    
    def _get_client_ip(self, request):
        """Get client IP address from request"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip
    
    def _check_rate_limit(self, ip):
        """Check if IP has exceeded rate limits"""
        
        hourly_key = f'contact_hourly_{ip}'
        daily_key = f'contact_daily_{ip}'
        
        hourly_count = cache.get(hourly_key, 0)
        daily_count = cache.get(daily_key, 0)
        
        if hourly_count >= self.RATE_LIMIT_HOURLY:
            return False, "Too many submissions. Please try again in an hour."
        
        if daily_count >= self.RATE_LIMIT_DAILY:
            return False, "Daily submission limit reached. Please try again tomorrow."
        
        # Increment counters
        cache.set(hourly_key, hourly_count + 1, 3600)  # 1 hour
        cache.set(daily_key, daily_count + 1, 86400)   # 24 hours
        
        return True, None
    
    def _is_spam_content(self, text):
        """Check if text contains spam patterns"""
        text_lower = text.lower()
        
        # Check for URLs (most spam contains links)
        if 'http://' in text_lower or 'https://' in text_lower or 'www.' in text_lower:
            return True, "Messages with URLs are not allowed."
        
        # Check for spam keywords
        for keyword in self.SPAM_KEYWORDS:
            if keyword in text_lower:
                return True, "Your message was flagged as potential spam."
        
        # Check for excessive caps (> 50% uppercase)
        if len(text) > 20:
            caps_count = sum(1 for c in text if c.isupper())
            if caps_count / len(text) > 0.5:
                return True, "Please don't use excessive capitalization."
        
        # Check message length
        if len(text.strip()) < self.MIN_MESSAGE_LENGTH:
            return True, "Please provide a more detailed message."
        
        return False, None
    
    def _is_disposable_email(self, email):
        """Check if email is from a disposable domain"""
        domain = email.split('@')[-1].lower()
        return domain in self.DISPOSABLE_DOMAINS

    def post(self, request):
        form = ContactForm(request.POST)
        
        # This handles validations for all fields including reCAPTCHA (if configured)
        if not form.is_valid():
            messages.error(request, "Please correct the errors below.")
            return render(request, self.template_name, {'form': form})

        # Get cleaned data
        data = form.cleaned_data
        name = data.get('name')
        email = data.get('email')
        subject = data.get('subject')
        message = data.get('message')
        honeypot = data.get('website')
        
        # Layer 1: Honeypot check
        if honeypot:
            # Silently reject spam bots - don't reveal honeypot was triggered
            messages.success(request, "Your message has been sent! We'll get back to you shortly.")
            return redirect('contact')
        
        # Layer 2: Rate limiting
        client_ip = self._get_client_ip(request)
        rate_ok, rate_msg = self._check_rate_limit(client_ip)
        if not rate_ok:
            messages.error(request, rate_msg)
            return render(request, self.template_name, {'form': form})
        
        # Layer 3: Content filtering
        is_spam, spam_msg = self._is_spam_content(subject + ' ' + message)
        if is_spam:
            messages.error(request, spam_msg)
            return render(request, self.template_name, {'form': form})
        
        # Layer 4: Email validation
        if self._is_disposable_email(email):
            messages.error(request, "Please use a permanent email address.")
            return render(request, self.template_name, {'form': form})
        
        # Layer 5: reCAPTCHA verification is handled by form.is_valid()
        
        # All checks passed - send email
        full_message = f"""
        New Contact Form Submission:
        
        Name: {name}
        Email: {email}
        Subject: {subject}
        IP: {client_ip}
        
        Message:
        {message}
        """

        try:
            send_mail(
                subject=f"Contact Form: {subject}",
                message=full_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=["track.my.rupee.app@gmail.com"],
                fail_silently=False,
            )
            messages.success(
                request, "Your message has been sent! We'll get back to you shortly."
            )
            return redirect("contact")
        except Exception as e:
            # Log error if possible
            messages.error(request, "Something went wrong. Please try again later.")
            return render(request, self.template_name, {"form": form})


@login_required
def predict_category_view(request):
    """
    AJAX view to predict category based on description.
    """
    if request.method == "GET":
        description = request.GET.get("description", "").strip()
        if not description:
            return JsonResponse({"category": None})

        category = predict_category_ai(description, request.user)
        return JsonResponse({"category": category})
    return JsonResponse({"error": "Invalid request"}, status=400)


# Friend Management Views (AJAX)
@login_required
def create_friend_ajax(request):
    """AJAX endpoint to create a new friend."""
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            name = data.get("name", "").strip()
            email = data.get("email", "").strip()
            phone = data.get("phone", "").strip()

            if not name:
                return JsonResponse({"success": False, "error": "Name is required"})

            # Check if friend already exists for this user
            if Friend.objects.filter(user=request.user, name=name).exists():
                return JsonResponse(
                    {
                        "success": False,
                        "error": "A friend with this name already exists",
                    }
                )

            # Create friend for current user
            friend = Friend.objects.create(
                user=request.user,
                name=name,
                email=email if email else None,
                phone=phone if phone else None,
            )

            return JsonResponse(
                {
                    "success": True,
                    "friend": {
                        "id": friend.id,
                        "name": friend.name,
                        "email": friend.email or "",
                        "phone": friend.phone or "",
                    },
                }
            )
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Invalid request method"})


@login_required
def update_friend_ajax(request, pk):
    """AJAX endpoint to update an existing friend."""
    if request.method == "POST":
        try:
            friend = get_object_or_404(Friend, pk=pk, user=request.user)
            data = json.loads(request.body)

            name = data.get("name", "").strip()
            email = data.get("email", "").strip()
            phone = data.get("phone", "").strip()

            if not name:
                return JsonResponse({"success": False, "error": "Name is required"})

            # Check if another friend of this user has this name
            if (
                Friend.objects.filter(user=request.user, name=name)
                .exclude(pk=pk)
                .exists()
            ):
                return JsonResponse(
                    {
                        "success": False,
                        "error": "A friend with this name already exists",
                    }
                )

            # Update friend
            friend.name = name
            friend.email = email if email else None
            friend.phone = phone if phone else None
            friend.save()

            return JsonResponse(
                {
                    "success": True,
                    "friend": {
                        "id": friend.id,
                        "name": friend.name,
                        "email": friend.email or "",
                        "phone": friend.phone or "",
                    },
                }
            )
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Invalid request method"})


# --------------------
# Notification Views
# --------------------

class NotificationListView(LoginRequiredMixin, ListView):
    model = Notification
    template_name = 'expenses/notification_list.html'
    context_object_name = 'notifications'
    paginate_by = 20

    def get_queryset(self):
        return Notification.objects.filter(user=self.request.user).order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['unread_count'] = Notification.objects.filter(user=self.request.user, is_read=False).count()
        return context

@login_required
def mark_notifications_read(request):
    if request.method == 'POST':
        Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
        messages.success(request, "All notifications marked as read.")
        return redirect('notification-list')
    return redirect('notification-list')

@login_required
def mark_single_notification_read(request, pk):
    try:
        notification = Notification.objects.get(pk=pk, user=request.user)
        notification.is_read = True
        notification.save()
        return JsonResponse({'success': True})
    except Notification.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Notification not found'}, status=404)

@csrf_exempt
def trigger_notifications(request):
    """
    HTTP endpoint to trigger notifications via external cron service (e.g. cron-job.org).
    Secured by a secret key in the URL params: ?secret=YOUR_CRON_SECRET
    """
    secret = request.GET.get('secret')

    # Check against dedicated CRON_SECRET
    if not secret or secret != settings.CRON_SECRET:
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    try:
        call_command('send_notifications')
        return JsonResponse({'success': True, 'message': 'Notifications triggered successfully'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    if request.method == 'POST':
        notification = get_object_or_404(Notification, pk=pk, user=request.user)
        notification.is_read = True
        notification.save()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False}, status=400)


class AnalyticsView(LoginRequiredMixin, TemplateView):
    template_name = 'expenses/analytics.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        today = timezone.now().date()

        # 1. Monthly Trends (Last 12 Months)
        labels = []
        income_data = []
        expense_data = []
        balance_rate_data = []

        # Determine the start date: 1st day of the month 11 months ago
        # If today is Jan 2026, 11 months ago is Feb 2025.
        start_date = (today.replace(day=1) - timedelta(days=365)).replace(day=1)

        # Fetch data grouped by Month
        monthly_income = Income.objects.filter(
            user=user, date__gte=start_date
        ).annotate(month=TruncMonth('date')).values('month').annotate(total=Sum('amount')).order_by('month')

        monthly_expenses = Expense.objects.filter(
            user=user, date__gte=start_date
        ).annotate(month=TruncMonth('date')).values('month').annotate(total=Sum('amount')).order_by('month')

        # Merge data into a map {date: {income: 0, expense: 0}}
        data_map = {}

        # Initialize map with all 12 months to ensure 0s for missing months
        # Iterate from start_date to today month by month
        curr = start_date
        while curr <= today:
            d = curr.replace(day=1)
            data_map[d] = {'income': 0, 'expense': 0}
            # Move to next month
            # Carefully handle month increment
            next_month = curr.month + 1
            next_year = curr.year
            if next_month > 12:
                next_month = 1
                next_year += 1
            curr = date(next_year, next_month, 1)

        # Fill with DB data
        # Fill with DB data
        for item in monthly_income:
            if item['month']:
                d = item['month']
                if isinstance(d, datetime):
                    d = d.date()
                d = d.replace(day=1)
                if d in data_map:
                    data_map[d]['income'] = float(item['total'])

        for item in monthly_expenses:
             if item['month']:
                d = item['month']
                if isinstance(d, datetime):
                    d = d.date()
                d = d.replace(day=1)
                if d in data_map:
                    data_map[d]['expense'] = float(item['total'])

        # Sort and prepare lists
        sorted_keys = sorted(data_map.keys())
        # Limit to last 12 months if while loop went over
        sorted_keys = sorted_keys[-12:]

        for k in sorted_keys:
            labels.append(k.strftime('%b %Y'))
            inc = data_map[k]['income']
            exp = data_map[k]['expense']
            income_data.append(inc)
            expense_data.append(exp)

            # Balance Rate = (Income - Expense) / Income * 100
            if inc > 0:
                rate = ((inc - exp) / inc) * 100
            else:
                rate = 0
            balance_rate_data.append(round(rate, 1))

        context['chart_labels'] = labels
        context['income_data'] = income_data
        context['expense_data'] = expense_data
        context['balance_rate_data'] = balance_rate_data

        # 2. Category Breakdown (Current Year)
        current_year = today.year
        category_stats = Expense.objects.filter(
            user=user, date__year=current_year
        ).values('category').annotate(total=Sum('amount')).order_by('-total')

        cat_labels = [x['category'] for x in category_stats]
        cat_data = [float(x['total']) for x in category_stats]

        context['cat_labels'] = cat_labels
        context['cat_data'] = cat_data

        # 3. Key Metrics (YTD)
        # Recalculate based on DB (more accurate than summing chart data if chart is limited)
        # Use date__lte=today to ensure we don't include future recurring entries or future dates
        ytd_income_agg = Income.objects.filter(user=user, date__year=current_year, date__lte=today).aggregate(Sum('amount'))['amount__sum'] or 0
        ytd_expense_agg = Expense.objects.filter(user=user, date__year=current_year, date__lte=today).aggregate(Sum('amount'))['amount__sum'] or 0

        context['total_income_ytd'] = ytd_income_agg
        context['total_expense_ytd'] = ytd_expense_agg
        context['total_balance_ytd'] = ytd_income_agg - ytd_expense_agg

        if ytd_income_agg > 0:
            context['avg_balance_rate'] = round(((ytd_income_agg - ytd_expense_agg) / ytd_income_agg) * 100, 1)
        else:
            context['avg_balance_rate'] = 0

        return context

@login_required
def delete_friend_ajax(request, pk):
    """AJAX endpoint to delete a friend."""
    if request.method == "POST":
        try:
            friend = get_object_or_404(Friend, pk=pk, user=request.user)
            friend_name = friend.name

            # Check if friend is used in any shared expenses
            if friend.expense_participations.exists():
                return JsonResponse(
                    {
                        "success": False,
                        "error": f"{friend_name} cannot be deleted because they are involved in shared expenses",
                    }
                )

            friend.delete()

            return JsonResponse(
                {"success": True, "message": f"{friend_name} deleted successfully"}
            )
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Invalid request method"})
